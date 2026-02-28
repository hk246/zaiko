"""Excel import service – template generation, parsing, validation, DB insertion."""

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session

from backend.models.material import RawMaterial
from backend.models.lot import Lot
from backend.models.reservation import Reservation
from backend.models.work_order import WorkOrder
from backend.models.master import Tare, Contact
from backend.models.label import MaterialLabel
from backend.models.property_field import PropertyField, LotProperty

# ── Column definitions per data type ──

COLUMNS: dict[str, list[dict[str, Any]]] = {
    "materials": [
        {"key": "name", "label": "名前", "required": True, "example": "酸化チタン"},
        {"key": "unit", "label": "単位", "required": False, "example": "g"},
        {"key": "min_weight", "label": "最低在庫量", "required": False, "example": "100"},
        {"key": "label_name", "label": "ラベル名", "required": False, "example": "顔料"},
    ],
    "lots": [
        {"key": "lot_name", "label": "ロット名", "required": True, "example": "LOT-2024-001"},
        {"key": "weight", "label": "重量", "required": True, "example": "500.0"},
        {"key": "is_fraction", "label": "端数フラグ", "required": False, "example": "FALSE"},
    ],
    "work_orders": [
        {"key": "process_name", "label": "工程名", "required": True, "example": "混合工程A"},
        {"key": "priority", "label": "優先度", "required": False, "example": "none"},
        {"key": "worker_name", "label": "担当者", "required": False, "example": "田中"},
        {"key": "experiment_type", "label": "種別", "required": False, "example": "standard"},
        {"key": "planned_start", "label": "予定開始日", "required": False, "example": "2025-04-01"},
        {"key": "planned_end", "label": "予定終了日", "required": False, "example": "2025-04-15"},
        {"key": "notes", "label": "メモ", "required": False, "example": ""},
    ],
    "reservations": [
        {"key": "material_name", "label": "材料名", "required": True, "example": "酸化チタン"},
        {"key": "type", "label": "種類", "required": True, "example": "use"},
        {"key": "quantity", "label": "数量", "required": True, "example": "50.0"},
        {"key": "scheduled_date", "label": "予定日", "required": True, "example": "2025-04-01"},
        {"key": "user_name", "label": "担当者", "required": False, "example": "田中"},
        {"key": "purpose", "label": "目的", "required": False, "example": "試作"},
    ],
    "tares": [
        {"key": "name", "label": "名前", "required": True, "example": "ビーカー100ml"},
        {"key": "weight", "label": "重量(g)", "required": True, "example": "85.5"},
        {"key": "description", "label": "説明", "required": False, "example": ""},
    ],
    "contacts": [
        {"key": "name", "label": "名前", "required": True, "example": "山田太郎"},
        {"key": "email", "label": "メール", "required": True, "example": "yamada@example.com"},
        {"key": "description", "label": "説明", "required": False, "example": ""},
    ],
    "labels": [
        {"key": "name", "label": "名前", "required": True, "example": "顔料"},
        {"key": "color", "label": "色コード", "required": False, "example": "#6c757d"},
    ],
    "property_fields": [
        {"key": "name", "label": "フィールド名", "required": True, "example": "粘度"},
        {"key": "field_type", "label": "型", "required": True, "example": "number"},
        {"key": "unit", "label": "単位", "required": False, "example": "mPa·s"},
        {"key": "label_name", "label": "ラベル名", "required": False, "example": "顔料"},
    ],
    "lot_properties": [
        {"key": "lot_name", "label": "ロット名", "required": True, "example": "LOT-2024-001"},
        {"key": "field_name", "label": "フィールド名", "required": True, "example": "粘度"},
        {"key": "value", "label": "値", "required": True, "example": "150.5"},
    ],
}

VALID_TYPES = set(COLUMNS.keys())

# Labels for data type names
TYPE_LABELS: dict[str, str] = {
    "materials": "材料",
    "lots": "ロット",
    "work_orders": "作業工程",
    "reservations": "予約",
    "tares": "風袋",
    "contacts": "連絡先",
    "labels": "ラベル",
    "property_fields": "物性値フィールド",
    "lot_properties": "物性値",
}


# ── Template Generation ──

def generate_template(data_type: str) -> BytesIO:
    """Generate an Excel template file for the given data type."""
    cols = COLUMNS[data_type]
    wb = Workbook()
    ws = wb.active
    ws.title = TYPE_LABELS.get(data_type, data_type)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write headers (row 1)
    for col_idx, col_def in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx)
        label = col_def["label"]
        if col_def["required"]:
            label += " *"
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(15, len(label) * 2 + 4)

    # Write example row (row 2)
    for col_idx, col_def in enumerate(cols, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = col_def["example"]
        if col_def["required"]:
            cell.fill = required_fill
        cell.border = thin_border

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Parsing ──

def _cell_value(cell) -> Any:
    """Extract value from an openpyxl cell, stripping strings."""
    v = cell.value
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def parse_excel(data_type: str, file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse an uploaded Excel file and return a list of row dicts."""
    cols = COLUMNS[data_type]
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Skip completely empty rows
        values = [_cell_value(c) for c in row[: len(cols)]]
        if all(v is None for v in values):
            continue
        row_dict: dict[str, Any] = {"_row": row_idx}
        for i, col_def in enumerate(cols):
            row_dict[col_def["key"]] = values[i] if i < len(values) else None
        rows.append(row_dict)
    return rows


# ── Validation ──

def _validate_required(row: dict, col_defs: list[dict], errors: list[dict], row_num: int):
    for col_def in col_defs:
        if col_def["required"] and row.get(col_def["key"]) is None:
            errors.append({"row": row_num, "field": col_def["label"], "message": "必須項目です"})


def _parse_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return date.fromisoformat(str(val).strip())
    except (ValueError, TypeError):
        return None


def _parse_bool(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("true", "1", "yes", "はい", "○")


def _get_label_name_map(db: Session) -> dict[str, int]:
    """Build a label name→id map from active labels."""
    labels = db.query(MaterialLabel).filter(MaterialLabel.deleted_at.is_(None)).all()
    return {l.name: l.id for l in labels}


def validate_materials(rows: list[dict], db: Session) -> list[dict]:
    errors: list[dict] = []
    label_map = _get_label_name_map(db)

    for row in rows:
        rn = row["_row"]
        _validate_required(row, COLUMNS["materials"], errors, rn)
        if row.get("min_weight") is not None:
            if _parse_float(row["min_weight"]) is None:
                errors.append({"row": rn, "field": "最低在庫量", "message": "数値を入力してください"})
        if row.get("label_name") is not None:
            if str(row["label_name"]) not in label_map:
                errors.append({"row": rn, "field": "ラベル名", "message": f"「{row['label_name']}」は存在しないラベルです"})
    return errors


def validate_lots(rows: list[dict], db: Session, material_id: int) -> list[dict]:
    errors: list[dict] = []
    # Check material exists
    mat = db.get(RawMaterial, material_id)
    if not mat or mat.deleted_at:
        return [{"row": 0, "field": "", "message": f"材料ID={material_id}が見つかりません"}]
    for row in rows:
        rn = row["_row"]
        _validate_required(row, COLUMNS["lots"], errors, rn)
        if row.get("weight") is not None:
            if _parse_float(row["weight"]) is None:
                errors.append({"row": rn, "field": "重量", "message": "数値を入力してください"})
    return errors


def validate_work_orders(rows: list[dict], db: Session) -> list[dict]:
    errors: list[dict] = []
    valid_priorities = {"critical", "high", "low", "none"}
    valid_types = {"standard", "experiment"}
    for row in rows:
        rn = row["_row"]
        _validate_required(row, COLUMNS["work_orders"], errors, rn)
        if row.get("priority") and str(row["priority"]).lower() not in valid_priorities:
            errors.append({"row": rn, "field": "優先度", "message": f"none/low/high/criticalのいずれかを指定してください"})
        if row.get("experiment_type") and str(row["experiment_type"]).lower() not in valid_types:
            errors.append({"row": rn, "field": "種別", "message": "standard/experimentのいずれかを指定してください"})
        for date_field, label in [("planned_start", "予定開始日"), ("planned_end", "予定終了日")]:
            if row.get(date_field) is not None and _parse_date(row[date_field]) is None:
                errors.append({"row": rn, "field": label, "message": "日付形式が不正です (YYYY-MM-DD)"})
    return errors


def validate_reservations(rows: list[dict], db: Session) -> list[dict]:
    errors: list[dict] = []
    # Preload material name map
    materials = db.query(RawMaterial).filter(RawMaterial.deleted_at.is_(None)).all()
    mat_name_map = {m.name: m.id for m in materials}

    valid_rsv_types = {"use", "replenish"}
    for row in rows:
        rn = row["_row"]
        _validate_required(row, COLUMNS["reservations"], errors, rn)
        if row.get("material_name") and str(row["material_name"]) not in mat_name_map:
            errors.append({"row": rn, "field": "材料名", "message": f"「{row['material_name']}」は存在しない材料です"})
        if row.get("type") and str(row["type"]).lower() not in valid_rsv_types:
            errors.append({"row": rn, "field": "種類", "message": "use/replenishのいずれかを指定してください"})
        if row.get("quantity") is not None:
            if _parse_float(row["quantity"]) is None:
                errors.append({"row": rn, "field": "数量", "message": "数値を入力してください"})
        if row.get("scheduled_date") is not None:
            if _parse_date(row["scheduled_date"]) is None:
                errors.append({"row": rn, "field": "予定日", "message": "日付形式が不正です (YYYY-MM-DD)"})
    return errors


def validate_tares(rows: list[dict], db: Session) -> list[dict]:
    errors: list[dict] = []
    for row in rows:
        rn = row["_row"]
        _validate_required(row, COLUMNS["tares"], errors, rn)
        if row.get("weight") is not None:
            if _parse_float(row["weight"]) is None:
                errors.append({"row": rn, "field": "重量(g)", "message": "数値を入力してください"})
    return errors


def validate_contacts(rows: list[dict], db: Session) -> list[dict]:
    errors: list[dict] = []
    for row in rows:
        rn = row["_row"]
        _validate_required(row, COLUMNS["contacts"], errors, rn)
        if row.get("email") is not None and "@" not in str(row.get("email", "")):
            errors.append({"row": rn, "field": "メール", "message": "有効なメールアドレスを入力してください"})
    return errors


def validate_labels(rows: list[dict], db: Session) -> list[dict]:
    errors: list[dict] = []
    existing = {l.name for l in db.query(MaterialLabel).filter(MaterialLabel.deleted_at.is_(None)).all()}
    seen: set[str] = set()
    for row in rows:
        rn = row["_row"]
        _validate_required(row, COLUMNS["labels"], errors, rn)
        name = row.get("name")
        if name is not None:
            name_str = str(name)
            if name_str in existing:
                errors.append({"row": rn, "field": "名前", "message": f"「{name_str}」は既に存在します"})
            if name_str in seen:
                errors.append({"row": rn, "field": "名前", "message": f"「{name_str}」がファイル内で重複しています"})
            seen.add(name_str)
        if row.get("color") is not None:
            c = str(row["color"]).strip()
            if not (c.startswith("#") and len(c) in (4, 7)):
                errors.append({"row": rn, "field": "色コード", "message": "#RRGGBB形式で入力してください"})
    return errors


def validate_property_fields(rows: list[dict], db: Session) -> list[dict]:
    errors: list[dict] = []
    label_map = _get_label_name_map(db)
    valid_field_types = {"string", "number"}
    for row in rows:
        rn = row["_row"]
        _validate_required(row, COLUMNS["property_fields"], errors, rn)
        if row.get("field_type") and str(row["field_type"]).lower() not in valid_field_types:
            errors.append({"row": rn, "field": "型", "message": "string/numberのいずれかを指定してください"})
        if row.get("label_name") is not None:
            if str(row["label_name"]) not in label_map:
                errors.append({"row": rn, "field": "ラベル名", "message": f"「{row['label_name']}」は存在しないラベルです"})
    return errors


def validate_lot_properties(rows: list[dict], db: Session, material_id: int) -> list[dict]:
    errors: list[dict] = []
    mat = db.get(RawMaterial, material_id)
    if not mat or mat.deleted_at:
        return [{"row": 0, "field": "", "message": f"材料ID={material_id}が見つかりません"}]

    # Build lot name→id map for this material
    lots = db.query(Lot).filter(Lot.material_id == material_id, Lot.deleted_at.is_(None)).all()
    lot_name_map = {l.lot_name: l.id for l in lots}

    # Build field name→field map (fields applicable to this material's label + shared fields)
    q = db.query(PropertyField)
    if mat.label_id:
        q = q.filter((PropertyField.label_id == mat.label_id) | (PropertyField.label_id.is_(None)))
    else:
        q = q.filter(PropertyField.label_id.is_(None))
    fields = q.all()
    field_name_map = {f.name: f for f in fields}

    for row in rows:
        rn = row["_row"]
        _validate_required(row, COLUMNS["lot_properties"], errors, rn)
        if row.get("lot_name") is not None:
            if str(row["lot_name"]) not in lot_name_map:
                errors.append({"row": rn, "field": "ロット名", "message": f"「{row['lot_name']}」はこの材料のロットに存在しません"})
        if row.get("field_name") is not None:
            fname = str(row["field_name"])
            if fname not in field_name_map:
                errors.append({"row": rn, "field": "フィールド名", "message": f"「{fname}」は利用可能なフィールドではありません"})
            else:
                field = field_name_map[fname]
                if field.field_type == "number" and row.get("value") is not None:
                    if _parse_float(row["value"]) is None:
                        errors.append({"row": rn, "field": "値", "message": f"「{fname}」は数値フィールドです。数値を入力してください"})
    return errors


VALIDATORS = {
    "materials": validate_materials,
    "lots": None,  # special: needs material_id
    "work_orders": validate_work_orders,
    "reservations": validate_reservations,
    "tares": validate_tares,
    "contacts": validate_contacts,
    "labels": validate_labels,
    "property_fields": validate_property_fields,
    "lot_properties": None,  # special: needs material_id
}


# ── Import (DB insertion) ──

def import_materials(rows: list[dict], db: Session) -> int:
    label_map = _get_label_name_map(db)
    count = 0
    for row in rows:
        label_id = label_map.get(str(row.get("label_name"))) if row.get("label_name") else None
        obj = RawMaterial(
            name=str(row["name"]),
            unit=str(row.get("unit") or "g"),
            min_weight=_parse_float(row.get("min_weight")) or 0.0,
            label_id=label_id,
        )
        db.add(obj)
        count += 1
    db.commit()
    return count


def import_lots(rows: list[dict], db: Session, material_id: int) -> int:
    count = 0
    for row in rows:
        obj = Lot(
            material_id=material_id,
            lot_name=str(row["lot_name"]),
            weight=_parse_float(row["weight"]) or 0.0,
            is_fraction=_parse_bool(row.get("is_fraction")),
        )
        db.add(obj)
        count += 1
    db.commit()
    return count


def import_work_orders(rows: list[dict], db: Session) -> int:
    count = 0
    for row in rows:
        obj = WorkOrder(
            process_name=str(row["process_name"]),
            priority=str(row.get("priority") or "none").lower(),
            worker_name=str(row["worker_name"]) if row.get("worker_name") else None,
            experiment_type=str(row.get("experiment_type") or "standard").lower(),
            planned_start=_parse_date(row.get("planned_start")),
            planned_end=_parse_date(row.get("planned_end")),
            notes=str(row["notes"]) if row.get("notes") else None,
        )
        db.add(obj)
        count += 1
    db.commit()
    return count


def import_reservations(rows: list[dict], db: Session) -> int:
    materials = db.query(RawMaterial).filter(RawMaterial.deleted_at.is_(None)).all()
    mat_name_map = {m.name: m.id for m in materials}
    count = 0
    for row in rows:
        material_id = mat_name_map[str(row["material_name"])]
        obj = Reservation(
            material_id=material_id,
            type=str(row["type"]).lower(),
            quantity=_parse_float(row["quantity"]) or 0.0,
            scheduled_date=_parse_date(row["scheduled_date"]),
            user_name=str(row["user_name"]) if row.get("user_name") else None,
            purpose=str(row["purpose"]) if row.get("purpose") else None,
        )
        db.add(obj)
        count += 1
    db.commit()
    return count


def import_tares(rows: list[dict], db: Session) -> int:
    count = 0
    for row in rows:
        obj = Tare(
            name=str(row["name"]),
            weight=_parse_float(row["weight"]) or 0.0,
            description=str(row["description"]) if row.get("description") else None,
        )
        db.add(obj)
        count += 1
    db.commit()
    return count


def import_contacts(rows: list[dict], db: Session) -> int:
    count = 0
    for row in rows:
        obj = Contact(
            name=str(row["name"]),
            email=str(row["email"]),
            description=str(row["description"]) if row.get("description") else None,
        )
        db.add(obj)
        count += 1
    db.commit()
    return count


def import_labels(rows: list[dict], db: Session) -> int:
    count = 0
    for row in rows:
        obj = MaterialLabel(
            name=str(row["name"]),
            color=str(row["color"]).strip() if row.get("color") else "#6c757d",
        )
        db.add(obj)
        count += 1
    db.commit()
    return count


def import_property_fields(rows: list[dict], db: Session) -> int:
    label_map = _get_label_name_map(db)
    # Get max order_index
    max_idx = db.query(PropertyField).count()
    count = 0
    for row in rows:
        label_id = label_map.get(str(row.get("label_name"))) if row.get("label_name") else None
        obj = PropertyField(
            name=str(row["name"]),
            field_type=str(row["field_type"]).lower(),
            unit=str(row["unit"]) if row.get("unit") else None,
            label_id=label_id,
            order_index=max_idx + count,
        )
        db.add(obj)
        count += 1
    db.commit()
    return count


def import_lot_properties(rows: list[dict], db: Session, material_id: int) -> int:
    mat = db.get(RawMaterial, material_id)
    lots = db.query(Lot).filter(Lot.material_id == material_id, Lot.deleted_at.is_(None)).all()
    lot_name_map = {l.lot_name: l.id for l in lots}

    q = db.query(PropertyField)
    if mat.label_id:
        q = q.filter((PropertyField.label_id == mat.label_id) | (PropertyField.label_id.is_(None)))
    else:
        q = q.filter(PropertyField.label_id.is_(None))
    fields = q.all()
    field_name_map = {f.name: f for f in fields}

    count = 0
    for row in rows:
        lot_id = lot_name_map[str(row["lot_name"])]
        field = field_name_map[str(row["field_name"])]
        # Upsert: check existing
        existing = db.query(LotProperty).filter(
            LotProperty.lot_id == lot_id, LotProperty.field_id == field.id
        ).first()
        val = row["value"]
        if existing:
            if field.field_type == "number":
                existing.value_number = _parse_float(val)
                existing.value_string = None
            else:
                existing.value_string = str(val) if val is not None else None
                existing.value_number = None
        else:
            prop = LotProperty(lot_id=lot_id, field_id=field.id)
            if field.field_type == "number":
                prop.value_number = _parse_float(val)
            else:
                prop.value_string = str(val) if val is not None else None
            db.add(prop)
        count += 1
    db.commit()
    return count


IMPORTERS = {
    "materials": import_materials,
    "lots": None,  # special: needs material_id
    "work_orders": import_work_orders,
    "reservations": import_reservations,
    "tares": import_tares,
    "contacts": import_contacts,
    "labels": import_labels,
    "property_fields": import_property_fields,
    "lot_properties": None,  # special: needs material_id
}
