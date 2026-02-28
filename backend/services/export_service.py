"""Export service – CSV, Excel, and full data export."""

import csv
import io
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from backend.models.lot import Lot
from backend.models.material import RawMaterial
from backend.models.property_field import LotProperty, PropertyField
from backend.models.reservation import Reservation
from backend.models.recipe import Recipe, RecipeItem
from backend.models.work_order import WorkOrder
from backend.models.stock_transaction import StockTransaction
from backend.models.label import MaterialLabel
from backend.models.master import Tare, Contact
from backend.services import stock_calculator


def export_materials_csv(db: Session) -> str:
    """Export all materials as CSV string."""
    materials = (
        db.query(RawMaterial)
        .filter(RawMaterial.deleted_at == None)  # noqa: E711
        .order_by(RawMaterial.name)
        .all()
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "名前", "単位", "現在在庫", "最低在庫", "予測在庫", "ラベル"])

    for m in materials:
        cs = stock_calculator.current_stock(db, m.id)
        ps = stock_calculator.predicted_stock(db, m.id)
        label_name = m.label.name if m.label else ""
        writer.writerow([m.id, m.name, m.unit, cs, m.min_weight, ps, label_name])

    return output.getvalue()


def export_properties_excel(db: Session, label_id: int | None = None) -> bytes:
    """Export lot properties as Excel bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "物性値"

    # Get fields
    query = db.query(PropertyField)
    if label_id:
        query = query.filter(PropertyField.label_id == label_id)
    fields = query.order_by(PropertyField.order_index).all()

    # Header
    headers = ["材料名", "ロット名"]
    for f in fields:
        header = f.name
        if f.unit:
            header += f" ({f.unit})"
        headers.append(header)
    ws.append(headers)

    # Rows: lots with their property values
    lots_query = db.query(Lot).filter(Lot.deleted_at == None)  # noqa: E711
    if label_id:
        lots_query = lots_query.join(RawMaterial).filter(RawMaterial.label_id == label_id)

    for lot in lots_query.order_by(Lot.material_id, Lot.created_at).all():
        row = [lot.material.name, lot.lot_name]
        props = {p.field_id: p for p in lot.properties}
        for f in fields:
            p = props.get(f.id)
            if p:
                row.append(p.value_number if f.field_type == "number" else p.value_string or "")
            else:
                row.append("")
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
# Full data export (multi-sheet workbook)
# ═══════════════════════════════════════════════════════════════

_HEADER_FONT = Font(name="Yu Gothic UI", bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_FONT = Font(name="Yu Gothic UI", size=10)
_CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")


def _style_sheet(ws, headers: list[str], col_widths: list[int] | None = None):
    """Apply header styling and column widths."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_row(ws, row_idx: int, values: list):
    """Write a data row with styling."""
    fill = _ALT_FILL if row_idx % 2 == 0 else None
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = _CELL_FONT
        cell.alignment = _CELL_ALIGN
        cell.border = _THIN_BORDER
        if fill:
            cell.fill = fill


def _fmt_dt(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    return dt.strftime("%Y-%m-%d %H:%M")


def _fmt_date(d) -> str | None:
    if d is None:
        return None
    return d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)


# ── Sheet builders ──

def _build_materials_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("材料")
    headers = ["ID", "名前", "単位", "最低在庫量", "ラベル", "担当連絡先", "連絡先メール",
               "アクション種別", "現在在庫", "端数在庫", "予測在庫", "低在庫", "ロット数", "作成日"]
    widths = [6, 20, 8, 12, 12, 14, 22, 14, 12, 12, 12, 10, 10, 18]
    _style_sheet(ws, headers, widths)

    materials = db.query(RawMaterial).filter(RawMaterial.deleted_at == None).order_by(RawMaterial.name).all()  # noqa: E711
    for i, m in enumerate(materials, 2):
        cs = stock_calculator.current_stock(db, m.id)
        fs = stock_calculator.fraction_stock(db, m.id)
        ps = stock_calculator.predicted_stock(db, m.id)
        low = stock_calculator.is_low_stock(db, m.id, m.min_weight)
        lot_count = len([lot for lot in m.lots if not lot.deleted_at])
        action_map = {"none": "なし", "email": "メール", "excel": "Excel"}
        _write_row(ws, i, [
            m.id, m.name, m.unit, m.min_weight,
            m.label.name if m.label else "",
            m.contact.name if m.contact else "",
            m.contact.email if m.contact else "",
            action_map.get(m.action_type, m.action_type),
            round(cs, 2), round(fs, 2), round(ps, 2),
            "○" if low else "",
            lot_count, _fmt_dt(m.created_at),
        ])


def _build_lots_merged_sheet(wb: Workbook, db: Session):
    """Build lots + properties merged sheet. Lot name is the primary key."""
    ws = wb.create_sheet("ロット＋物性値")

    # Gather all property fields (ordered)
    fields = db.query(PropertyField).order_by(PropertyField.label_id, PropertyField.order_index).all()

    # Base headers + dynamic property columns
    base_headers = ["ID", "材料名", "ラベル", "ロット名", "重量(g)", "端数", "作成日"]
    prop_headers = []
    for f in fields:
        col_name = f.name
        if f.unit:
            col_name += f" ({f.unit})"
        prop_headers.append(col_name)

    headers = base_headers + prop_headers
    widths = [6, 20, 12, 18, 12, 8, 18] + [14] * len(prop_headers)
    _style_sheet(ws, headers, widths)

    # Field ID → column index mapping
    field_id_to_col = {f.id: idx for idx, f in enumerate(fields)}

    # Query lots
    lots = (
        db.query(Lot)
        .filter(Lot.deleted_at == None)  # noqa: E711
        .join(RawMaterial)
        .filter(RawMaterial.deleted_at == None)  # noqa: E711
        .order_by(RawMaterial.name, Lot.lot_name)
        .all()
    )

    for i, lot in enumerate(lots, 2):
        base_values = [
            lot.id,
            lot.material.name if lot.material else "",
            lot.material.label.name if (lot.material and lot.material.label) else "",
            lot.lot_name,
            round(lot.weight, 2),
            "○" if lot.is_fraction else "",
            _fmt_dt(lot.created_at),
        ]
        # Fill property values
        prop_values = [None] * len(fields)
        for lp in lot.properties:
            col = field_id_to_col.get(lp.field_id)
            if col is not None:
                if lp.field.field_type == "number" and lp.value_number is not None:
                    prop_values[col] = lp.value_number
                else:
                    prop_values[col] = lp.value_string
        _write_row(ws, i, base_values + prop_values)


def _build_reservations_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("予約")
    headers = ["ID", "材料名", "種別", "数量", "実績数量", "ロット名",
               "担当者", "目的", "予定日", "ステータス", "実行日時", "作成日"]
    widths = [6, 20, 10, 12, 12, 16, 12, 20, 14, 12, 18, 18]
    _style_sheet(ws, headers, widths)

    reservations = (
        db.query(Reservation)
        .join(RawMaterial)
        .filter(RawMaterial.deleted_at == None)  # noqa: E711
        .order_by(Reservation.scheduled_date.desc())
        .all()
    )
    type_map = {"use": "使用", "replenish": "補充"}
    status_map = {"pending": "未実行", "executed": "実行済", "cancelled": "キャンセル"}

    for i, r in enumerate(reservations, 2):
        _write_row(ws, i, [
            r.id,
            r.material.name if r.material else "",
            type_map.get(r.type, r.type),
            round(r.quantity, 2),
            round(r.actual_quantity, 2) if r.actual_quantity is not None else "",
            r.lot_name or "",
            r.user_name or "",
            r.purpose or "",
            _fmt_date(r.scheduled_date),
            status_map.get(r.status, r.status),
            _fmt_dt(r.executed_at),
            _fmt_dt(r.created_at),
        ])


def _build_work_orders_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("作業工程")
    headers = ["ID", "工程名", "ロット名", "材料名", "ステータス", "優先度",
               "担当者", "種別", "伝票発行", "予定開始", "予定終了",
               "実績開始", "実績終了", "メモ", "作成日"]
    widths = [6, 24, 16, 16, 12, 10, 12, 12, 10, 14, 14, 14, 14, 24, 18]
    _style_sheet(ws, headers, widths)

    orders = db.query(WorkOrder).filter(WorkOrder.archived == False).order_by(WorkOrder.created_at.desc()).all()  # noqa: E712
    status_map = {"pending": "未着手", "in_progress": "進行中", "completed": "完了"}
    priority_map = {"critical": "最重要", "high": "高", "low": "低", "none": "なし"}
    type_map = {"standard": "通常", "experiment": "実験"}

    for i, wo in enumerate(orders, 2):
        _write_row(ws, i, [
            wo.id, wo.process_name, wo.lot_name or "",
            wo.material.name if wo.material else "",
            status_map.get(wo.status, wo.status),
            priority_map.get(wo.priority, wo.priority),
            wo.worker_name or "",
            type_map.get(wo.experiment_type, wo.experiment_type),
            "○" if wo.invoice_issued else "",
            _fmt_date(wo.planned_start), _fmt_date(wo.planned_end),
            _fmt_date(wo.actual_start), _fmt_date(wo.actual_end),
            wo.notes or "", _fmt_dt(wo.created_at),
        ])


def _build_recipes_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("レシピ")
    headers = ["レシピID", "レシピ名", "種別", "説明", "材料名", "数量", "ロット名"]
    widths = [10, 20, 10, 24, 20, 12, 16]
    _style_sheet(ws, headers, widths)

    recipes = db.query(Recipe).filter(Recipe.deleted_at == None).order_by(Recipe.name).all()  # noqa: E711
    type_map = {"use": "使用", "replenish": "補充"}

    row = 2
    for recipe in recipes:
        items = db.query(RecipeItem).filter(RecipeItem.recipe_id == recipe.id).all()
        if not items:
            _write_row(ws, row, [
                recipe.id, recipe.name, type_map.get(recipe.type, recipe.type),
                recipe.description or "", "", "", "",
            ])
            row += 1
        else:
            for j, item in enumerate(items):
                _write_row(ws, row, [
                    recipe.id if j == 0 else "",
                    recipe.name if j == 0 else "",
                    type_map.get(recipe.type, recipe.type) if j == 0 else "",
                    (recipe.description or "") if j == 0 else "",
                    item.material.name if item.material else "",
                    round(item.quantity, 2),
                    item.lot_name or "",
                ])
                row += 1


def _build_labels_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("ラベル")
    headers = ["ID", "名前", "色", "作成日"]
    widths = [6, 20, 10, 18]
    _style_sheet(ws, headers, widths)

    labels = db.query(MaterialLabel).filter(MaterialLabel.deleted_at == None).order_by(MaterialLabel.name).all()  # noqa: E711
    for i, lb in enumerate(labels, 2):
        _write_row(ws, i, [lb.id, lb.name, lb.color, _fmt_dt(lb.created_at)])


def _build_tares_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("風袋")
    headers = ["ID", "名前", "重量(g)", "説明", "作成日"]
    widths = [6, 20, 12, 24, 18]
    _style_sheet(ws, headers, widths)

    tares = db.query(Tare).order_by(Tare.name).all()
    for i, t in enumerate(tares, 2):
        _write_row(ws, i, [t.id, t.name, round(t.weight, 2), t.description or "", _fmt_dt(t.created_at)])


def _build_contacts_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("連絡先")
    headers = ["ID", "名前", "メール", "説明", "作成日"]
    widths = [6, 20, 24, 24, 18]
    _style_sheet(ws, headers, widths)

    contacts = db.query(Contact).order_by(Contact.name).all()
    for i, c in enumerate(contacts, 2):
        _write_row(ws, i, [c.id, c.name, c.email, c.description or "", _fmt_dt(c.created_at)])


def _build_transactions_sheet(wb: Workbook, db: Session):
    ws = wb.create_sheet("在庫履歴")
    headers = ["ID", "材料名", "ロット名", "種別", "数量", "実行者", "メモ", "実行日時"]
    widths = [6, 20, 16, 10, 12, 12, 24, 18]
    _style_sheet(ws, headers, widths)

    txns = db.query(StockTransaction).order_by(StockTransaction.executed_at.desc()).all()
    type_map = {"use": "使用", "replenish": "補充"}

    for i, t in enumerate(txns, 2):
        _write_row(ws, i, [
            t.id,
            t.material.name if t.material else "",
            t.lot.lot_name if t.lot else "",
            type_map.get(t.type, t.type),
            round(t.quantity, 2),
            t.executed_by or "",
            t.note or "",
            _fmt_dt(t.executed_at),
        ])


def generate_export_workbook(db: Session) -> BytesIO:
    """Generate an Excel workbook with all data across multiple sheets."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    _build_materials_sheet(wb, db)
    _build_lots_merged_sheet(wb, db)
    _build_reservations_sheet(wb, db)
    _build_work_orders_sheet(wb, db)
    _build_recipes_sheet(wb, db)
    _build_labels_sheet(wb, db)
    _build_tares_sheet(wb, db)
    _build_contacts_sheet(wb, db)
    _build_transactions_sheet(wb, db)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
