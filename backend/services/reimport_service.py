"""Reimport service – parse exported Excel and upsert into DB."""

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.models.material import RawMaterial
from backend.models.lot import Lot
from backend.models.property_field import PropertyField, LotProperty
from backend.models.reservation import Reservation
from backend.models.work_order import WorkOrder
from backend.models.label import MaterialLabel
from backend.models.master import Tare, Contact


# ── Reverse maps (Japanese → DB values) ──

_TYPE_REV = {"使用": "use", "補充": "replenish"}
_RSV_STATUS_REV = {"未実行": "pending", "実行済": "executed", "キャンセル": "cancelled"}
_WO_STATUS_REV = {"未着手": "pending", "進行中": "in_progress", "完了": "completed"}
_PRIORITY_REV = {"最重要": "critical", "高": "high", "低": "low", "なし": "none"}
_EXP_TYPE_REV = {"通常": "standard", "実験": "experiment"}
_ACTION_REV = {"なし": "none", "メール": "email", "Excel": "excel"}


def _cell(cell) -> Any:
    v = cell.value
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def _parse_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_int(val: Any) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _parse_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return date.fromisoformat(str(val).strip()[:10])
    except (ValueError, TypeError):
        return None


def _parse_bool_marker(val: Any) -> bool:
    """Parse ○ or similar markers as True."""
    if val is None:
        return False
    return str(val).strip() in ("○", "TRUE", "true", "1", "はい")


def _rev(mapping: dict, val: Any, default: str = "") -> str:
    """Reverse map Japanese label to DB value."""
    if val is None:
        return default
    s = str(val).strip()
    return mapping.get(s, s)


# ── Sheet parsers ──

def _read_sheet_rows(ws) -> list[dict[str, Any]]:
    """Read a worksheet into list of {header: value} dicts."""
    headers = [_cell(c) for c in ws[1]]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [_cell(c) for c in row[:len(headers)]]
        if all(v is None for v in values):
            continue
        d = {"_row": row_idx}
        for i, h in enumerate(headers):
            if h is not None:
                d[h] = values[i] if i < len(values) else None
        rows.append(d)
    return rows


# ── Validators ──

def _validate_materials(rows: list[dict], db: Session,
                        pending_labels: set[str] | None = None,
                        pending_contacts: set[str] | None = None) -> tuple[list[dict], int, int]:
    errors = []
    updates = creates = 0
    label_map = {l.name: l.id for l in db.query(MaterialLabel).filter(MaterialLabel.deleted_at.is_(None)).all()}
    contact_map = {c.name: c.id for c in db.query(Contact).all()}
    existing_names = {m.name for m in db.query(RawMaterial).filter(RawMaterial.deleted_at.is_(None)).all()}
    # Include pending creates from earlier sheets
    all_labels = set(label_map.keys()) | (pending_labels or set())
    all_contacts = set(contact_map.keys()) | (pending_contacts or set())

    for row in rows:
        rn = row["_row"]
        rid = _parse_int(row.get("ID"))
        if rid:
            m = db.get(RawMaterial, rid)
            if not m or m.deleted_at:
                errors.append({"row": rn, "field": "ID", "message": f"ID={rid}の材料が見つかりません"})
                continue
            updates += 1
        else:
            if not row.get("名前"):
                errors.append({"row": rn, "field": "名前", "message": "必須項目です"})
                continue
            if str(row["名前"]) in existing_names:
                updates += 1
            else:
                creates += 1
                existing_names.add(str(row["名前"]))

        if row.get("最低在庫量") is not None and _parse_float(row["最低在庫量"]) is None:
            errors.append({"row": rn, "field": "最低在庫量", "message": "数値を入力してください"})
        if row.get("ラベル") and str(row["ラベル"]) not in all_labels:
            errors.append({"row": rn, "field": "ラベル", "message": f"「{row['ラベル']}」は存在しないラベルです"})
        if row.get("担当連絡先") and str(row["担当連絡先"]) not in all_contacts:
            errors.append({"row": rn, "field": "担当連絡先", "message": f"「{row['担当連絡先']}」は存在しない連絡先です"})

    return errors, updates, creates


def _validate_lots(rows: list[dict], db: Session, prop_fields: list[PropertyField],
                   pending_materials: set[str] | None = None) -> tuple[list[dict], int, int]:
    errors = []
    updates = creates = 0
    mat_map = {m.name: m.id for m in db.query(RawMaterial).filter(RawMaterial.deleted_at.is_(None)).all()}
    all_materials = set(mat_map.keys()) | (pending_materials or set())

    # Build set of existing (material_id, lot_name) pairs for name-based matching
    existing_lots = set()
    for lot in db.query(Lot).filter(Lot.deleted_at.is_(None)).all():
        existing_lots.add((lot.material_id, lot.lot_name))

    for row in rows:
        rn = row["_row"]
        rid = _parse_int(row.get("ID"))
        if rid:
            lot = db.get(Lot, rid)
            if not lot or lot.deleted_at:
                errors.append({"row": rn, "field": "ID", "message": f"ID={rid}のロットが見つかりません"})
                continue
            updates += 1
        else:
            if not row.get("ロット名"):
                errors.append({"row": rn, "field": "ロット名", "message": "必須項目です"})
                continue
            if not row.get("材料名") or str(row["材料名"]) not in all_materials:
                errors.append({"row": rn, "field": "材料名", "message": "有効な材料名を指定してください"})
                continue
            # Name-based: check if (material_id, lot_name) already exists
            mat_id = mat_map.get(str(row["材料名"]))
            lot_key = (mat_id, str(row["ロット名"])) if mat_id else None
            if lot_key and lot_key in existing_lots:
                updates += 1
            else:
                creates += 1
                if lot_key:
                    existing_lots.add(lot_key)

        if row.get("重量(g)") is not None and _parse_float(row["重量(g)"]) is None:
            errors.append({"row": rn, "field": "重量(g)", "message": "数値を入力してください"})

        # Validate property values
        for f in prop_fields:
            col_name = f.name + (f" ({f.unit})" if f.unit else "")
            val = row.get(col_name)
            if val is not None and f.field_type == "number" and _parse_float(val) is None:
                errors.append({"row": rn, "field": col_name, "message": "数値を入力してください"})

    return errors, updates, creates


def _validate_reservations(rows: list[dict], db: Session,
                           pending_materials: set[str] | None = None) -> tuple[list[dict], int, int]:
    errors = []
    updates = creates = 0
    mat_map = {m.name: m.id for m in db.query(RawMaterial).filter(RawMaterial.deleted_at.is_(None)).all()}
    all_materials = set(mat_map.keys()) | (pending_materials or set())

    for row in rows:
        rn = row["_row"]
        rid = _parse_int(row.get("ID"))
        if rid:
            r = db.get(Reservation, rid)
            if not r:
                errors.append({"row": rn, "field": "ID", "message": f"ID={rid}の予約が見つかりません"})
                continue
            updates += 1
        else:
            for field in ["材料名", "種別", "数量", "予定日"]:
                if not row.get(field):
                    errors.append({"row": rn, "field": field, "message": "必須項目です"})
            creates += 1

        if row.get("材料名") and str(row["材料名"]) not in all_materials:
            errors.append({"row": rn, "field": "材料名", "message": f"「{row['材料名']}」は存在しない材料です"})
        if row.get("数量") is not None and _parse_float(row["数量"]) is None:
            errors.append({"row": rn, "field": "数量", "message": "数値を入力してください"})
        if row.get("予定日") is not None and _parse_date(row["予定日"]) is None:
            errors.append({"row": rn, "field": "予定日", "message": "日付形式が不正です"})

    return errors, updates, creates


def _validate_work_orders(rows: list[dict], db: Session) -> tuple[list[dict], int, int]:
    errors = []
    updates = creates = 0

    for row in rows:
        rn = row["_row"]
        rid = _parse_int(row.get("ID"))
        if rid:
            wo = db.get(WorkOrder, rid)
            if not wo:
                errors.append({"row": rn, "field": "ID", "message": f"ID={rid}の作業工程が見つかりません"})
                continue
            updates += 1
        else:
            if not row.get("工程名"):
                errors.append({"row": rn, "field": "工程名", "message": "必須項目です"})
            creates += 1

        for date_field in ["予定開始", "予定終了", "実績開始", "実績終了"]:
            if row.get(date_field) is not None and _parse_date(row[date_field]) is None:
                errors.append({"row": rn, "field": date_field, "message": "日付形式が不正です"})

    return errors, updates, creates


def _validate_labels(rows: list[dict], db: Session) -> tuple[list[dict], int, int]:
    errors = []
    updates = creates = 0
    existing_names = {l.name for l in db.query(MaterialLabel).filter(MaterialLabel.deleted_at.is_(None)).all()}

    for row in rows:
        rn = row["_row"]
        rid = _parse_int(row.get("ID"))
        if rid:
            lb = db.get(MaterialLabel, rid)
            if not lb or lb.deleted_at:
                errors.append({"row": rn, "field": "ID", "message": f"ID={rid}のラベルが見つかりません"})
                continue
            updates += 1
        else:
            if not row.get("名前"):
                errors.append({"row": rn, "field": "名前", "message": "必須項目です"})
                continue
            # Name-based: existing name → update, new name → create
            if str(row["名前"]) in existing_names:
                updates += 1
            else:
                creates += 1
                existing_names.add(str(row["名前"]))  # track for duplicate detection

        if row.get("色") is not None:
            c = str(row["色"]).strip()
            if not (c.startswith("#") and len(c) in (4, 7)):
                errors.append({"row": rn, "field": "色", "message": "#RRGGBB形式で入力してください"})

    return errors, updates, creates


def _validate_tares(rows: list[dict], db: Session) -> tuple[list[dict], int, int]:
    errors = []
    updates = creates = 0
    existing_names = {t.name for t in db.query(Tare).all()}

    for row in rows:
        rn = row["_row"]
        rid = _parse_int(row.get("ID"))
        if rid:
            t = db.get(Tare, rid)
            if not t:
                errors.append({"row": rn, "field": "ID", "message": f"ID={rid}の風袋が見つかりません"})
                continue
            updates += 1
        else:
            if not row.get("名前"):
                errors.append({"row": rn, "field": "名前", "message": "必須項目です"})
                continue
            if str(row["名前"]) in existing_names:
                updates += 1
            else:
                creates += 1
                existing_names.add(str(row["名前"]))

        if row.get("重量(g)") is not None and _parse_float(row["重量(g)"]) is None:
            errors.append({"row": rn, "field": "重量(g)", "message": "数値を入力してください"})

    return errors, updates, creates


def _validate_contacts(rows: list[dict], db: Session) -> tuple[list[dict], int, int]:
    errors = []
    updates = creates = 0
    existing_names = {c.name for c in db.query(Contact).all()}

    for row in rows:
        rn = row["_row"]
        rid = _parse_int(row.get("ID"))
        if rid:
            c = db.get(Contact, rid)
            if not c:
                errors.append({"row": rn, "field": "ID", "message": f"ID={rid}の連絡先が見つかりません"})
                continue
            updates += 1
        else:
            for field in ["名前", "メール"]:
                if not row.get(field):
                    errors.append({"row": rn, "field": field, "message": "必須項目です"})
            if row.get("名前") and str(row["名前"]) in existing_names:
                updates += 1
            else:
                creates += 1
                if row.get("名前"):
                    existing_names.add(str(row["名前"]))

        if row.get("メール") is not None and "@" not in str(row.get("メール", "")):
            errors.append({"row": rn, "field": "メール", "message": "有効なメールアドレスを入力してください"})

    return errors, updates, creates


# ── Upsert functions ──

def _upsert_materials(rows: list[dict], db: Session) -> tuple[int, int]:
    label_map = {l.name: l.id for l in db.query(MaterialLabel).filter(MaterialLabel.deleted_at.is_(None)).all()}
    contact_map = {c.name: c.id for c in db.query(Contact).all()}
    updates = creates = 0

    for row in rows:
        rid = _parse_int(row.get("ID"))
        label_id = label_map.get(str(row.get("ラベル"))) if row.get("ラベル") else None
        contact_id = contact_map.get(str(row.get("担当連絡先"))) if row.get("担当連絡先") else None
        action = _rev(_ACTION_REV, row.get("アクション種別"), "none")

        m = None
        if rid:
            m = db.get(RawMaterial, rid)
            if not m or m.deleted_at:
                continue
        elif row.get("名前"):
            # Name-based lookup
            m = db.query(RawMaterial).filter(
                RawMaterial.name == str(row["名前"]),
                RawMaterial.deleted_at.is_(None),
            ).first()

        if m:
            if row.get("名前") is not None:
                m.name = str(row["名前"])
            if row.get("単位") is not None:
                m.unit = str(row["単位"])
            if row.get("最低在庫量") is not None:
                m.min_weight = _parse_float(row["最低在庫量"]) or 0.0
            m.label_id = label_id
            m.contact_id = contact_id
            m.action_type = action
            updates += 1
        else:
            m = RawMaterial(
                name=str(row.get("名前", "")),
                unit=str(row.get("単位") or "g"),
                min_weight=_parse_float(row.get("最低在庫量")) or 0.0,
                label_id=label_id,
                contact_id=contact_id,
                action_type=action,
            )
            db.add(m)
            creates += 1

    db.flush()
    return updates, creates


def _upsert_lots(rows: list[dict], db: Session, prop_fields: list[PropertyField]) -> tuple[int, int]:
    mat_map = {m.name: m.id for m in db.query(RawMaterial).filter(RawMaterial.deleted_at.is_(None)).all()}
    field_col_map = {}
    for f in prop_fields:
        col_name = f.name + (f" ({f.unit})" if f.unit else "")
        field_col_map[col_name] = f

    updates = creates = 0

    for row in rows:
        rid = _parse_int(row.get("ID"))
        lot = None

        if rid:
            lot = db.get(Lot, rid)
            if not lot or lot.deleted_at:
                continue
        elif row.get("ロット名") and row.get("材料名"):
            # Name-based lookup: match by (material_id, lot_name)
            material_id = mat_map.get(str(row["材料名"]))
            if material_id:
                lot = db.query(Lot).filter(
                    Lot.material_id == material_id,
                    Lot.lot_name == str(row["ロット名"]),
                    Lot.deleted_at.is_(None),
                ).first()

        if lot:
            if row.get("ロット名") is not None:
                lot.lot_name = str(row["ロット名"])
            if row.get("重量(g)") is not None:
                lot.weight = _parse_float(row["重量(g)"]) or 0.0
            lot.is_fraction = _parse_bool_marker(row.get("端数"))
            updates += 1
        else:
            material_id = mat_map.get(str(row.get("材料名")))
            if not material_id:
                continue
            lot = Lot(
                material_id=material_id,
                lot_name=str(row.get("ロット名", "")),
                weight=_parse_float(row.get("重量(g)")) or 0.0,
                is_fraction=_parse_bool_marker(row.get("端数")),
            )
            db.add(lot)
            db.flush()  # get lot.id
            creates += 1

        # Update property values
        for col_name, field in field_col_map.items():
            val = row.get(col_name)
            if val is None:
                continue
            existing = db.query(LotProperty).filter(
                LotProperty.lot_id == lot.id, LotProperty.field_id == field.id
            ).first()
            if existing:
                if field.field_type == "number":
                    existing.value_number = _parse_float(val)
                    existing.value_string = None
                else:
                    existing.value_string = str(val)
                    existing.value_number = None
            else:
                prop = LotProperty(lot_id=lot.id, field_id=field.id)
                if field.field_type == "number":
                    prop.value_number = _parse_float(val)
                else:
                    prop.value_string = str(val)
                db.add(prop)

    db.flush()
    return updates, creates


def _upsert_reservations(rows: list[dict], db: Session) -> tuple[int, int]:
    mat_map = {m.name: m.id for m in db.query(RawMaterial).filter(RawMaterial.deleted_at.is_(None)).all()}
    updates = creates = 0

    for row in rows:
        rid = _parse_int(row.get("ID"))
        material_id = mat_map.get(str(row.get("材料名"))) if row.get("材料名") else None
        rsv_type = _rev(_TYPE_REV, row.get("種別"), "use")
        status = _rev(_RSV_STATUS_REV, row.get("ステータス"), "pending")

        if rid:
            r = db.get(Reservation, rid)
            if not r:
                continue
            if material_id:
                r.material_id = material_id
            r.type = rsv_type
            if row.get("数量") is not None:
                r.quantity = _parse_float(row["数量"]) or 0.0
            if row.get("実績数量") is not None:
                r.actual_quantity = _parse_float(row["実績数量"])
            r.lot_name = str(row["ロット名"]) if row.get("ロット名") else r.lot_name
            r.user_name = str(row["担当者"]) if row.get("担当者") else r.user_name
            r.purpose = str(row["目的"]) if row.get("目的") else r.purpose
            if row.get("予定日") is not None:
                r.scheduled_date = _parse_date(row["予定日"]) or r.scheduled_date
            r.status = status
            updates += 1
        else:
            if not material_id:
                continue
            r = Reservation(
                material_id=material_id,
                type=rsv_type,
                quantity=_parse_float(row.get("数量")) or 0.0,
                actual_quantity=_parse_float(row.get("実績数量")),
                lot_name=str(row["ロット名"]) if row.get("ロット名") else None,
                user_name=str(row["担当者"]) if row.get("担当者") else None,
                purpose=str(row["目的"]) if row.get("目的") else None,
                scheduled_date=_parse_date(row.get("予定日")) or date.today(),
                status=status,
            )
            db.add(r)
            creates += 1

    db.flush()
    return updates, creates


def _upsert_work_orders(rows: list[dict], db: Session) -> tuple[int, int]:
    mat_map = {m.name: m.id for m in db.query(RawMaterial).filter(RawMaterial.deleted_at.is_(None)).all()}
    updates = creates = 0

    for row in rows:
        rid = _parse_int(row.get("ID"))
        status = _rev(_WO_STATUS_REV, row.get("ステータス"), "pending")
        priority = _rev(_PRIORITY_REV, row.get("優先度"), "none")
        exp_type = _rev(_EXP_TYPE_REV, row.get("種別"), "standard")
        material_id = mat_map.get(str(row.get("材料名"))) if row.get("材料名") else None

        if rid:
            wo = db.get(WorkOrder, rid)
            if not wo:
                continue
            if row.get("工程名") is not None:
                wo.process_name = str(row["工程名"])
            wo.lot_name = str(row["ロット名"]) if row.get("ロット名") else wo.lot_name
            wo.material_id = material_id if material_id else wo.material_id
            wo.status = status
            wo.priority = priority
            wo.worker_name = str(row["担当者"]) if row.get("担当者") else wo.worker_name
            wo.experiment_type = exp_type
            wo.invoice_issued = _parse_bool_marker(row.get("伝票発行"))
            wo.planned_start = _parse_date(row.get("予定開始")) if row.get("予定開始") is not None else wo.planned_start
            wo.planned_end = _parse_date(row.get("予定終了")) if row.get("予定終了") is not None else wo.planned_end
            wo.actual_start = _parse_date(row.get("実績開始")) if row.get("実績開始") is not None else wo.actual_start
            wo.actual_end = _parse_date(row.get("実績終了")) if row.get("実績終了") is not None else wo.actual_end
            wo.notes = str(row["メモ"]) if row.get("メモ") else wo.notes
            updates += 1
        else:
            wo = WorkOrder(
                process_name=str(row.get("工程名", "")),
                lot_name=str(row["ロット名"]) if row.get("ロット名") else None,
                material_id=material_id,
                status=status,
                priority=priority,
                worker_name=str(row["担当者"]) if row.get("担当者") else None,
                experiment_type=exp_type,
                invoice_issued=_parse_bool_marker(row.get("伝票発行")),
                planned_start=_parse_date(row.get("予定開始")),
                planned_end=_parse_date(row.get("予定終了")),
                actual_start=_parse_date(row.get("実績開始")),
                actual_end=_parse_date(row.get("実績終了")),
                notes=str(row["メモ"]) if row.get("メモ") else None,
            )
            db.add(wo)
            creates += 1

    db.flush()
    return updates, creates


def _upsert_labels(rows: list[dict], db: Session) -> tuple[int, int]:
    updates = creates = 0
    for row in rows:
        rid = _parse_int(row.get("ID"))
        lb = None

        if rid:
            lb = db.get(MaterialLabel, rid)
            if not lb or lb.deleted_at:
                continue
        elif row.get("名前"):
            # Name-based lookup: if label with same name exists, update it
            lb = db.query(MaterialLabel).filter(
                MaterialLabel.name == str(row["名前"]),
                MaterialLabel.deleted_at.is_(None),
            ).first()

        if lb:
            if row.get("名前") is not None:
                lb.name = str(row["名前"])
            if row.get("色") is not None:
                lb.color = str(row["色"]).strip()
            updates += 1
        else:
            lb = MaterialLabel(
                name=str(row.get("名前", "")),
                color=str(row["色"]).strip() if row.get("色") else "#6c757d",
            )
            db.add(lb)
            creates += 1
    db.flush()
    return updates, creates


def _upsert_tares(rows: list[dict], db: Session) -> tuple[int, int]:
    updates = creates = 0
    for row in rows:
        rid = _parse_int(row.get("ID"))
        t = None

        if rid:
            t = db.get(Tare, rid)
            if not t:
                continue
        elif row.get("名前"):
            # Name-based lookup
            t = db.query(Tare).filter(Tare.name == str(row["名前"])).first()

        if t:
            if row.get("名前") is not None:
                t.name = str(row["名前"])
            if row.get("重量(g)") is not None:
                t.weight = _parse_float(row["重量(g)"]) or 0.0
            if row.get("説明") is not None:
                t.description = str(row["説明"]) or None
            updates += 1
        else:
            t = Tare(
                name=str(row.get("名前", "")),
                weight=_parse_float(row.get("重量(g)")) or 0.0,
                description=str(row["説明"]) if row.get("説明") else None,
            )
            db.add(t)
            creates += 1
    db.flush()
    return updates, creates


def _upsert_contacts(rows: list[dict], db: Session) -> tuple[int, int]:
    updates = creates = 0
    for row in rows:
        rid = _parse_int(row.get("ID"))
        c = None

        if rid:
            c = db.get(Contact, rid)
            if not c:
                continue
        elif row.get("名前"):
            # Name-based lookup
            c = db.query(Contact).filter(Contact.name == str(row["名前"])).first()

        if c:
            if row.get("名前") is not None:
                c.name = str(row["名前"])
            if row.get("メール") is not None:
                c.email = str(row["メール"])
            if row.get("説明") is not None:
                c.description = str(row["説明"]) or None
            updates += 1
        else:
            c = Contact(
                name=str(row.get("名前", "")),
                email=str(row.get("メール", "")),
                description=str(row["説明"]) if row.get("説明") else None,
            )
            db.add(c)
            creates += 1
    db.flush()
    return updates, creates


# ── Sheet name to handler mapping ──

_SHEET_NAMES = ["材料", "ロット＋物性値", "予約", "作業工程", "ラベル", "風袋", "連絡先"]
_SKIP_SHEETS = {"レシピ", "在庫履歴"}


# ── Main functions ──

def _collect_pending_names(rows: list[dict], name_field: str = "名前") -> set[str]:
    """Collect all names from sheet rows (both updates and creates).

    This is used for cross-sheet dependency validation: e.g. when the labels
    sheet updates label ID=1 from name 'a' to '原料A', the materials sheet
    should accept '原料A' as a valid label name even though it doesn't exist
    in the DB yet.
    """
    names = set()
    for row in rows:
        if row.get(name_field):
            names.add(str(row[name_field]))
    return names


def preview_reimport(file_bytes: bytes, db: Session) -> dict:
    """Parse and validate all sheets in dependency order, return preview summary."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    prop_fields = db.query(PropertyField).order_by(PropertyField.label_id, PropertyField.order_index).all()

    # Parse all sheets upfront
    sheet_rows: dict[str, list[dict]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in _SKIP_SHEETS:
            continue
        if sheet_name not in _SHEET_NAMES:
            continue
        ws = wb[sheet_name]
        sheet_rows[sheet_name] = _read_sheet_rows(ws)

    sheets_result = []
    all_valid = True

    # Collect pending names from earlier sheets for dependency resolution
    pending_labels: set[str] = set()
    pending_contacts: set[str] = set()
    pending_materials: set[str] = set()

    if "ラベル" in sheet_rows:
        pending_labels = _collect_pending_names(sheet_rows["ラベル"])
    if "連絡先" in sheet_rows:
        pending_contacts = _collect_pending_names(sheet_rows["連絡先"])
    if "材料" in sheet_rows:
        pending_materials = _collect_pending_names(sheet_rows["材料"])

    # Validate in dependency order
    process_order = ["ラベル", "連絡先", "風袋", "材料", "ロット＋物性値", "予約", "作業工程"]

    for sheet_name in process_order:
        if sheet_name not in sheet_rows:
            continue
        rows = sheet_rows[sheet_name]
        if not rows:
            sheets_result.append({"name": sheet_name, "total": 0, "updates": 0, "creates": 0, "errors": []})
            continue

        if sheet_name == "ラベル":
            errors, updates, creates = _validate_labels(rows, db)
        elif sheet_name == "連絡先":
            errors, updates, creates = _validate_contacts(rows, db)
        elif sheet_name == "風袋":
            errors, updates, creates = _validate_tares(rows, db)
        elif sheet_name == "材料":
            errors, updates, creates = _validate_materials(rows, db, pending_labels, pending_contacts)
        elif sheet_name == "ロット＋物性値":
            errors, updates, creates = _validate_lots(rows, db, prop_fields, pending_materials)
        elif sheet_name == "予約":
            errors, updates, creates = _validate_reservations(rows, db, pending_materials)
        elif sheet_name == "作業工程":
            errors, updates, creates = _validate_work_orders(rows, db)
        else:
            continue

        if errors:
            all_valid = False

        sheets_result.append({
            "name": sheet_name,
            "total": len(rows),
            "updates": updates,
            "creates": creates,
            "errors": errors,
        })

    return {"sheets": sheets_result, "valid": all_valid}


def execute_reimport(file_bytes: bytes, db: Session) -> dict:
    """Parse, validate (dependency-aware), and upsert all sheets.

    Validates and upserts each sheet in dependency order so that
    earlier sheets (labels, contacts) are committed before later
    sheets (materials, lots) that reference them.
    """
    # First do a dependency-aware preview
    preview = preview_reimport(file_bytes, db)
    if not preview["valid"]:
        raise ValueError("バリデーションエラーがあります。プレビューで確認してください。")

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    prop_fields = db.query(PropertyField).order_by(PropertyField.label_id, PropertyField.order_index).all()

    total_updates = 0
    total_creates = 0
    sheet_results = []

    # Process in order: labels first (materials depend on them), then materials, lots, etc.
    process_order = ["ラベル", "連絡先", "風袋", "材料", "ロット＋物性値", "予約", "作業工程"]

    for sheet_name in process_order:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = _read_sheet_rows(ws)
        if not rows:
            continue

        if sheet_name == "ラベル":
            u, c = _upsert_labels(rows, db)
        elif sheet_name == "連絡先":
            u, c = _upsert_contacts(rows, db)
        elif sheet_name == "風袋":
            u, c = _upsert_tares(rows, db)
        elif sheet_name == "材料":
            u, c = _upsert_materials(rows, db)
        elif sheet_name == "ロット＋物性値":
            u, c = _upsert_lots(rows, db, prop_fields)
        elif sheet_name == "予約":
            u, c = _upsert_reservations(rows, db)
        elif sheet_name == "作業工程":
            u, c = _upsert_work_orders(rows, db)
        else:
            continue

        total_updates += u
        total_creates += c
        sheet_results.append({"name": sheet_name, "updates": u, "creates": c})

    db.commit()
    return {
        "total_updates": total_updates,
        "total_creates": total_creates,
        "sheets": sheet_results,
    }
