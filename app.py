from flask import Flask, render_template, request, redirect, url_for, make_response, jsonify, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect
from wtforms import StringField, FloatField, SubmitField, SelectField, DateField
from wtforms.validators import DataRequired, Email, Optional
from datetime import datetime, date
import sys
import csv
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import shutil
import os
import json
from pathlib import Path
from tkinter import Tk, filedialog, messagebox
import webbrowser
import threading
import subprocess
import platform

# ======================================================
# 実行ファイル化（Nuitka/PyInstaller）対応: ベースディレクトリ解決
# ======================================================
def _get_base_dir():
    """実行環境に応じてアプリのベースディレクトリを返す"""
    # PyInstaller onefile/standalone
    if getattr(sys, 'frozen', False):
        return os.path.dirname(os.path.realpath(sys.executable))
    # Nuitka standalone / onefile
    try:
        _ = __compiled__   # Nuitka コンパイル時にのみ定義される
        return os.path.dirname(os.path.realpath(sys.executable))
    except NameError:
        pass
    # 通常の Python 実行
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = _get_base_dir()

# 設定ファイルのパス（常に実行ファイルと同じディレクトリ）
CONFIG_FILE = os.path.join(BASE_DIR, 'config.json')

def load_config():
    """設定ファイルを読み込む"""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_config(config):
    """設定ファイルに保存"""
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

def select_database_folder():
    """データベースフォルダを選択"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    folder = filedialog.askdirectory(
        title='データベースフォルダを選択してください',
        initialdir=BASE_DIR
    )
    
    root.destroy()
    return folder

def get_database_path():
    """データベースパスを取得または設定"""
    config = load_config()
    
    # 設定にデータベースフォルダがあるか確認
    if 'database_folder' in config and os.path.exists(config['database_folder']):
        db_folder = config['database_folder']
    else:
        # フォルダ選択ダイアログを表示
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        messagebox.showinfo(
            'データベースフォルダの選択',
            'データベースを保存するフォルダを選択してください。\n'
            '共有フォルダを指定すると、複数人で同じデータベースを使用できます。'
        )
        
        db_folder = select_database_folder()
        root.destroy()
        
        if not db_folder:
            messagebox.showerror('エラー', 'フォルダが選択されませんでした。\nデフォルトのinstanceフォルダを使用します。')
            db_folder = os.path.join(BASE_DIR, 'instance')
            os.makedirs(db_folder, exist_ok=True)
        
        # 設定を保存
        config['database_folder'] = db_folder
        save_config(config)
    
    # データベースファイルのパスを返す
    db_path = os.path.join(db_folder, 'inventory.db')
    return db_path

# Flaskアプリケーションの初期化
# template_folder/static_folderを絶対パスで指定（実行ファイル化対応）
app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, 'templates'),
    static_folder=os.path.join(BASE_DIR, 'static'),
)
app.config['SECRET_KEY'] = 'your-secret-key'

# データベースパスを取得
db_path = get_database_path()
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'

db = SQLAlchemy(app)
csrf = CSRFProtect(app)

class MaterialLabel(db.Model):
    """原料ラベル（カテゴリ）"""
    __tablename__ = 'material_label'
    id           = db.Column(db.Integer, primary_key=True)
    name         = db.Column(db.String(50), nullable=False, unique=True)
    color        = db.Column(db.String(7), nullable=False, default='#6c757d')
    date_created = db.Column(db.DateTime, default=db.func.current_timestamp())

    def __repr__(self):
        return f'<MaterialLabel {self.name}>'


class RawMaterial(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    weight = db.Column(db.Float, nullable=False)  # 原料全体の重量（表示用）
    unit = db.Column(db.String(20), default='g')  # 単位はg固定
    min_weight = db.Column(db.Float, default=0.0)
    email = db.Column(db.String(120), nullable=True)  # 購入担当者メール
    excel_path = db.Column(db.String(500), nullable=True)  # エクセルファイルパス
    action_type = db.Column(db.String(20), default='none')  # 'email', 'excel', 'none'
    label_id = db.Column(db.Integer, db.ForeignKey('material_label.id'), nullable=True)
    label    = db.relationship('MaterialLabel', backref=db.backref('materials', lazy=True))

    def get_total_lot_weight(self):
        """全ロットの現在重量の合計（端数を除く）"""
        return sum(lot.weight for lot in self.lots if not getattr(lot, 'is_fraction', False))
    
    def get_fraction_lot_weight(self):
        """端数ロットの重量合計"""
        return sum(lot.weight for lot in self.lots if getattr(lot, 'is_fraction', False))
    
    def get_predicted_stock(self):
        """現在量 + 未実行の補充予約 - 未実行の使用予約 = 予測在庫量（ロットの合計）"""
        total_current = self.get_total_lot_weight()
        # 実行済みの予約はカウントしない
        replenish = sum(r.quantity for r in self.reservations if r.type == 'replenish' and not r.executed)
        use = sum(r.quantity for r in self.reservations if r.type == 'use' and not r.executed)
        return total_current + replenish - use
    
    def is_low_stock_alert(self):
        """予測在庫が最低量を下回るかチェック（途中の期間も含む）"""
        # 途中で最低量を下回る期間がある場合もアラートとする
        critical_periods = self.get_critical_periods()
        return len(critical_periods) > 0

    def get_critical_periods(self):
        """最低重量を下回る期間を計算"""
        from datetime import datetime, timedelta
        
        # 現在の在庫量
        current_stock = self.get_total_lot_weight()
        
        # 未実行の予約を日付順に取得
        reservations = sorted(
            [r for r in self.reservations if not r.executed and r.scheduled_date],
            key=lambda x: x.scheduled_date
        )
        
        if not reservations:
            # 予約がない場合、現在の在庫が最低重量を下回っているかチェック
            if current_stock < self.min_weight:
                return [{
                    'start_date': datetime.now().date(),
                    'end_date': None,
                    'min_stock': current_stock,
                    'shortage': self.min_weight - current_stock
                }]
            return []
        
        critical_periods = []
        running_stock = current_stock
        period_start = None
        period_start_date = None
        min_stock_in_period = running_stock
        
        # 現在の在庫が既に不足している場合
        if running_stock < self.min_weight:
            period_start = True
            period_start_date = datetime.now().date()
            min_stock_in_period = running_stock
        
        # 各予約を時系列で処理
        for reservation in reservations:
            # 予約実行前の在庫状態をチェック
            prev_stock = running_stock
            
            # 予約を実行
            if reservation.type == 'use':
                running_stock -= reservation.quantity
            else:  # replenish
                running_stock += reservation.quantity
            
            # 使用予約で最低重量を下回った場合、期間開始
            if reservation.type == 'use' and prev_stock >= self.min_weight and running_stock < self.min_weight:
                period_start = True
                period_start_date = reservation.scheduled_date
                min_stock_in_period = running_stock
            
            # 既に期間中で、さらに在庫が減少
            elif period_start and running_stock < self.min_weight:
                min_stock_in_period = min(min_stock_in_period, running_stock)
            
            # 補充予約で最低重量を上回った場合、期間終了
            if reservation.type == 'replenish' and period_start and running_stock >= self.min_weight:
                critical_periods.append({
                    'start_date': period_start_date,
                    'end_date': reservation.scheduled_date,
                    'min_stock': min_stock_in_period,
                    'shortage': self.min_weight - min_stock_in_period
                })
                period_start = False
                period_start_date = None
                min_stock_in_period = running_stock
        
        # 最後の期間が終了していない場合
        if period_start:
            critical_periods.append({
                'start_date': period_start_date,
                'end_date': None,  # 終了日未定（補充予約が必要）
                'min_stock': min_stock_in_period,
                'shortage': self.min_weight - min_stock_in_period
            })
        
        return critical_periods

    def get_usage_stats(self, period_days):
        """指定期間の使用量・補充量を集計"""
        from datetime import datetime, timedelta
        
        end_date = datetime.now()
        start_date = end_date - timedelta(days=period_days)
        
        # 実行済みの予約のみを対象（実行日時が記録されているもの）
        executed_reservations = [r for r in self.reservations if r.executed and r.executed_date]
        
        # 期間内の予約をフィルタ
        period_reservations = [
            r for r in executed_reservations 
            if start_date <= r.executed_date <= end_date
        ]
        
        # 使用量と補充量を集計
        total_used = sum(r.actual_quantity or r.quantity for r in period_reservations if r.type == 'use')
        total_replenished = sum(r.actual_quantity or r.quantity for r in period_reservations if r.type == 'replenish')
        
        # 日別データ
        daily_data = {}
        for r in period_reservations:
            date_key = r.executed_date.strftime('%Y-%m-%d')
            if date_key not in daily_data:
                daily_data[date_key] = {'used': 0, 'replenished': 0}
            
            if r.type == 'use':
                daily_data[date_key]['used'] += r.actual_quantity or r.quantity
            else:
                daily_data[date_key]['replenished'] += r.actual_quantity or r.quantity
        
        return {
            'period_days': period_days,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'total_used': round(total_used, 3),
            'total_replenished': round(total_replenished, 3),
            'net_change': round(total_replenished - total_used, 3),
            'daily_data': daily_data,
            'transaction_count': len(period_reservations)
        }

    def __repr__(self):
        return f'<RawMaterial {self.name}>'

class Lot(db.Model):
    """ロット（原料の下位管理単位）"""
    id = db.Column(db.Integer, primary_key=True)
    material_id = db.Column(db.Integer, db.ForeignKey('raw_material.id'), nullable=False)
    lot_name = db.Column(db.String(100), nullable=False)  # ロット名
    weight = db.Column(db.Float, nullable=False)  # ロットの重量
    is_fraction = db.Column(db.Boolean, default=False)  # 端数処理フラグ
    date_created = db.Column(db.DateTime, default=db.func.current_timestamp())

    material = db.relationship('RawMaterial', backref=db.backref('lots', lazy=True, cascade='all, delete-orphan'))

    def __repr__(self):
        return f'<Lot {self.lot_name}>'

class Reservation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    material_id = db.Column(db.Integer, db.ForeignKey('raw_material.id'), nullable=False)
    lot_id = db.Column(db.Integer, db.ForeignKey('lot.id'), nullable=True)  # 既存ロット指定（オプショナル）
    lot_name = db.Column(db.String(100), nullable=True)  # 新規ロット名（オプショナル）
    recipe_id = db.Column(db.Integer, db.ForeignKey('recipe.id'), nullable=True)  # レシピからの予約
    type = db.Column(db.String(20), nullable=False)  # 'use' or 'replenish'
    quantity = db.Column(db.Float, nullable=False)  # 予約量
    actual_quantity = db.Column(db.Float, nullable=True)  # 実際の量
    user_name = db.Column(db.String(100), nullable=True)  # 使用者名
    purpose = db.Column(db.String(200), nullable=True)  # 目的
    scheduled_date = db.Column(db.Date, nullable=True)  # 予定日
    date = db.Column(db.DateTime, default=db.func.current_timestamp())  # 登録日
    executed = db.Column(db.Boolean, default=False)  # 実行済みかどうか
    executed_date = db.Column(db.DateTime, nullable=True)  # 実行日時

    material = db.relationship('RawMaterial', backref=db.backref('reservations', lazy=True, cascade='all, delete-orphan'))
    lot = db.relationship('Lot', backref=db.backref('reservations', lazy=True, cascade='all, delete-orphan'))
    recipe = db.relationship('Recipe', backref=db.backref('reservations', lazy=True))

    def is_overdue(self):
        """期限切れかどうかをチェック"""
        if not self.scheduled_date or self.executed:
            return False
        return self.scheduled_date < date.today()

    def __repr__(self):
        return f'<Reservation {self.type} {self.quantity}>'

class ReservationUsage(db.Model):
    """予約実行時の複数ロット使用記録"""
    id = db.Column(db.Integer, primary_key=True)
    reservation_id = db.Column(db.Integer, db.ForeignKey('reservation.id'), nullable=False)
    lot_id = db.Column(db.Integer, db.ForeignKey('lot.id'), nullable=False)
    quantity = db.Column(db.Float, nullable=False)  # このロットから使用した量
    date_created = db.Column(db.DateTime, default=db.func.current_timestamp())

    reservation = db.relationship('Reservation', backref=db.backref('usages', lazy=True, cascade='all, delete-orphan'))
    lot = db.relationship('Lot', backref=db.backref('usages', lazy=True))

    def __repr__(self):
        return f'<ReservationUsage reservation={self.reservation_id} lot={self.lot_id} quantity={self.quantity}>'

class Recipe(db.Model):
    """複数原料の組み合わせ（レシピ）"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # レシピ名
    description = db.Column(db.String(200), nullable=True)  # 説明
    type = db.Column(db.String(20), nullable=False)  # 'use' or 'replenish'
    date_created = db.Column(db.DateTime, default=db.func.current_timestamp())

    def __repr__(self):
        return f'<Recipe {self.name}>'

class RecipeItem(db.Model):
    """レシピの各原料と量"""
    id = db.Column(db.Integer, primary_key=True)
    recipe_id = db.Column(db.Integer, db.ForeignKey('recipe.id'), nullable=False)
    material_id = db.Column(db.Integer, db.ForeignKey('raw_material.id'), nullable=False)
    quantity = db.Column(db.Float, nullable=False)
    lot_name = db.Column(db.String(100), nullable=True)  # ロット名（オプショナル）

    recipe = db.relationship('Recipe', backref=db.backref('items', lazy=True, cascade='all, delete-orphan'))
    material = db.relationship('RawMaterial')

    def __repr__(self):
        return f'<RecipeItem {self.material.name} {self.quantity}>'

class Tare(db.Model):
    """風袋（容器）の管理"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # 風袋名
    weight = db.Column(db.Float, nullable=False)  # 風袋重量（g）
    description = db.Column(db.String(200), nullable=True)  # 説明
    date_created = db.Column(db.DateTime, default=db.func.current_timestamp())

    def __repr__(self):
        return f'<Tare {self.name} {self.weight}g>'

class Contact(db.Model):
    """連絡先（アドレス帳）"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # 名前
    email = db.Column(db.String(120), nullable=False)  # メールアドレス
    description = db.Column(db.String(200), nullable=True)  # 備考
    date_created = db.Column(db.DateTime, default=db.func.current_timestamp())

    def __repr__(self):
        return f'<Contact {self.name} {self.email}>'

class WorkOrder(db.Model):
    """作業工程進捗管理"""
    id = db.Column(db.Integer, primary_key=True)
    process_name = db.Column(db.String(200), nullable=False)       # 作業工程名
    lot_name = db.Column(db.String(100), nullable=True)            # 製造ロット名
    material_id = db.Column(db.Integer, db.ForeignKey('raw_material.id'), nullable=True)  # 関連原料（任意）
    status = db.Column(db.String(20), default='pending')           # pending / in_progress / completed
    priority = db.Column(db.String(20), default='none')            # critical / high / low / none
    invoice_issued = db.Column(db.Boolean, default=False)          # 伝票発行済
    worker_name = db.Column(db.String(100), nullable=True)          # 作業者名
    experiment_type = db.Column(db.String(20), default='standard') # standard / experiment
    planned_start = db.Column(db.Date, nullable=True)              # 開始予定日
    planned_end = db.Column(db.Date, nullable=True)                # 完了予定日
    actual_start = db.Column(db.Date, nullable=True)               # 開始日
    actual_end = db.Column(db.Date, nullable=True)                 # 完了日
    notes = db.Column(db.Text, nullable=True)                      # 備考
    archived = db.Column(db.Boolean, default=False)                # アーカイブ済み
    date_created = db.Column(db.DateTime, default=db.func.current_timestamp())

    material = db.relationship('RawMaterial', backref=db.backref('work_orders', lazy=True))

    STATUS_LABELS = {
        'pending': '取り掛かり前',
        'in_progress': '作業中',
        'completed': '完了'
    }
    PRIORITY_LABELS = {
        'critical': '最優先',
        'high': '優先大',
        'low': '優先小',
        'none': '未入力'
    }
    PRIORITY_ORDER = {'critical': 0, 'high': 1, 'low': 2, 'none': 3}

    def status_label(self):
        return self.STATUS_LABELS.get(self.status, self.status)

    def priority_label(self):
        return self.PRIORITY_LABELS.get(self.priority, self.priority)

    def compute_status(self):
        """実際の日付からステータスを自動判定
        - actual_start 未入力       → pending（取り掛かり前）
        - actual_start 入力済み     → in_progress（作業中）
        - actual_end が今日以前     → completed（完了）
        """
        if self.actual_end and self.actual_end <= date.today():
            return 'completed'
        elif self.actual_start:
            return 'in_progress'
        else:
            return 'pending'

    def is_overdue(self):
        """完了予定日を超えていて未完了の場合"""
        if self.planned_end and self.status != 'completed':
            return self.planned_end < date.today()
        return False

    def progress_pct(self):
        """開始〜完了予定日に対する経過率（0〜100）"""
        if not self.planned_start or not self.planned_end:
            return 0
        total = (self.planned_end - self.planned_start).days
        if total <= 0:
            return 100
        elapsed = (date.today() - self.planned_start).days
        return max(0, min(100, int(elapsed / total * 100)))

    def __repr__(self):
        return f'<WorkOrder {self.process_name}>'


class PropertyField(db.Model):
    """物性値フィールド定義（ラベルごと）"""
    __tablename__ = 'property_field'
    id           = db.Column(db.Integer, primary_key=True)
    label_id     = db.Column(db.Integer, db.ForeignKey('material_label.id'), nullable=True)
    name         = db.Column(db.String(100), nullable=False)
    field_type   = db.Column(db.String(20), nullable=False, default='string')  # 'string' or 'number'
    unit         = db.Column(db.String(30), nullable=True)
    order_index  = db.Column(db.Integer, default=0)
    date_created = db.Column(db.DateTime, default=db.func.current_timestamp())

    label = db.relationship('MaterialLabel',
                backref=db.backref('property_fields', lazy=True,
                                   order_by='PropertyField.order_index',
                                   cascade='all, delete-orphan'))

    def __repr__(self):
        return f'<PropertyField {self.name}>'


class LotProperty(db.Model):
    """ロット物性値"""
    __tablename__ = 'lot_property'
    id           = db.Column(db.Integer, primary_key=True)
    lot_id       = db.Column(db.Integer, db.ForeignKey('lot.id'), nullable=False)
    field_id     = db.Column(db.Integer, db.ForeignKey('property_field.id'), nullable=False)
    value_string = db.Column(db.String(500), nullable=True)
    value_number = db.Column(db.Float, nullable=True)
    date_updated = db.Column(db.DateTime, default=db.func.current_timestamp())

    lot   = db.relationship('Lot', backref=db.backref('properties', lazy=True,
                                                      cascade='all, delete-orphan'))
    field = db.relationship('PropertyField',
                backref=db.backref('lot_properties', lazy=True))

    __table_args__ = (db.UniqueConstraint('lot_id', 'field_id'),)

    def __repr__(self):
        return f'<LotProperty lot={self.lot_id} field={self.field_id}>'


class MaterialForm(FlaskForm):
    name = StringField('Name', validators=[DataRequired()])
    weight = FloatField('Weight (g)', validators=[Optional()], default=0.0)
    min_weight = FloatField('Min Weight (g)', default=0.0)
    action_type = SelectField('Action Type', choices=[('none', 'なにもしない'), ('email', 'メール連絡'), ('excel', 'エクセルを開く')], default='none')
    email = StringField('Purchase Email', validators=[Optional(), Email()])
    excel_path = StringField('Excel File Path', validators=[Optional()])
    submit = SubmitField('Submit')

class LotForm(FlaskForm):
    lot_name = StringField('Lot Name', validators=[DataRequired()])
    use_tare = SelectField('Use Tare', coerce=int, validators=[Optional()])
    weight = FloatField('Weight', validators=[DataRequired()])
    submit = SubmitField('Submit')

class ReservationForm(FlaskForm):
    lot_id = SelectField('Existing Lot (Optional)', coerce=int, validators=[Optional()])
    lot_name = StringField('New Lot Name (Optional)', validators=[Optional()])
    quantity = FloatField('Quantity', validators=[DataRequired()])

    user_name = StringField('User Name (Optional)', validators=[Optional()])
    purpose = StringField('Purpose (Optional)', validators=[Optional()])
    scheduled_date = DateField('Scheduled Date (Optional)', format='%Y-%m-%d', validators=[Optional()])
    submit = SubmitField('Reserve')

class RecipeForm(FlaskForm):
    name = StringField('Recipe Name', validators=[DataRequired()])
    description = StringField('Description (Optional)', validators=[Optional()])
    submit = SubmitField('Save Recipe')

class TareForm(FlaskForm):
    name = StringField('Tare Name', validators=[DataRequired()])
    weight = FloatField('Weight (g)', validators=[DataRequired()])
    description = StringField('Description (Optional)', validators=[Optional()])
    submit = SubmitField('Save')

class ContactForm(FlaskForm):
    name = StringField('Name', validators=[DataRequired()])
    email = StringField('Email', validators=[DataRequired(), Email()])
    description = StringField('Description (Optional)', validators=[Optional()])
    submit = SubmitField('Save')

@app.route('/')
def index():
    search   = request.args.get('search', '')
    sort_by  = request.args.get('sort_by', 'name')
    label_id = request.args.get('label_id', '')
    labels   = MaterialLabel.query.order_by(MaterialLabel.name).all()
    materials = RawMaterial.query
    if search:
        materials = materials.filter(RawMaterial.name.contains(search))
    if label_id == 'none':
        materials = materials.filter(RawMaterial.label_id == None)
    elif label_id:
        try:
            materials = materials.filter(RawMaterial.label_id == int(label_id))
        except ValueError:
            pass
    if sort_by == 'name':
        materials = materials.order_by(RawMaterial.name)
    elif sort_by == 'weight':
        # ソート用: SQLでは直接計算できないため、Pythonでソート
        materials = materials.all()
        materials = sorted(materials, key=lambda m: m.get_total_lot_weight())
        return render_template('index.html', materials=materials, search=search,
                               sort_by=sort_by, label_id=label_id, labels=labels)
    materials = materials.all()
    return render_template('index.html', materials=materials, search=search,
                           sort_by=sort_by, label_id=label_id, labels=labels)

@app.route('/add', methods=['GET', 'POST'])
def add():
    form = MaterialForm()
    contacts = Contact.query.order_by(Contact.name).all()
    labels   = MaterialLabel.query.order_by(MaterialLabel.name).all()
    if form.validate_on_submit():
        label_id_raw = request.form.get('label_id', '')
        label_id_val = int(label_id_raw) if label_id_raw else None
        material = RawMaterial(
            name=form.name.data,
            weight=form.weight.data if form.weight.data is not None else 0.0,
            unit='g',  # g固定
            min_weight=form.min_weight.data,
            email=form.email.data,
            excel_path=form.excel_path.data,
            action_type=form.action_type.data,
            label_id=label_id_val
        )
        db.session.add(material)
        db.session.commit()
        return redirect(url_for('index'))
    return render_template('add.html', form=form, contacts=contacts, labels=labels)

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit(id):
    material = RawMaterial.query.get_or_404(id)
    form = MaterialForm()
    contacts = Contact.query.order_by(Contact.name).all()
    labels   = MaterialLabel.query.order_by(MaterialLabel.name).all()
    if form.validate_on_submit():
        label_id_raw      = request.form.get('label_id', '')
        material.label_id = int(label_id_raw) if label_id_raw else None
        material.name = form.name.data
        material.weight = form.weight.data
        material.unit = 'g'  # g固定
        material.min_weight = form.min_weight.data
        material.email = form.email.data
        material.excel_path = form.excel_path.data
        material.action_type = form.action_type.data
        db.session.commit()
        return redirect(url_for('index'))
    elif request.method == 'GET':
        form.name.data = material.name
        form.weight.data = material.weight
        form.min_weight.data = material.min_weight
        form.email.data = material.email
        form.excel_path.data = material.excel_path
        form.action_type.data = material.action_type
    return render_template('edit.html', form=form, contacts=contacts, labels=labels, material=material)

@app.route('/material_stats/<int:id>')
def material_stats(id):
    """原料の統計ページ"""
    material = RawMaterial.query.get_or_404(id)
    return render_template('material_stats.html', material=material)

@app.route('/delete/<int:id>')
def delete(id):
    material = RawMaterial.query.get_or_404(id)
    name = material.name

    try:
        # ロットIDを先に取得
        lot_ids = [lot.id for lot in material.lots]

        # このロットを参照するReservationUsageを削除
        if lot_ids:
            ReservationUsage.query.filter(ReservationUsage.lot_id.in_(lot_ids)).delete(synchronize_session=False)

        # この原料の予約を参照するReservationUsageを削除（lot_idがNoneのもの含む）
        reservation_ids = [r.id for r in material.reservations]
        if reservation_ids:
            ReservationUsage.query.filter(ReservationUsage.reservation_id.in_(reservation_ids)).delete(synchronize_session=False)

        # 原料を削除（cascade='all, delete-orphan' で予約・ロットも連鎖削除）
        db.session.delete(material)
        db.session.commit()
        flash(f'原料「{name}」と関連データを削除しました', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'削除に失敗しました: {str(e)}', 'danger')

    return redirect(url_for('index'))

@app.route('/reserve_use/<int:id>', methods=['GET', 'POST'])
def reserve_use(id):
    material = RawMaterial.query.get_or_404(id)
    form = ReservationForm()
    # ロット選択肢を追加（端数ロットは除外）
    form.lot_id.choices = [(0, '既存ロットから選択しない')] + [(lot.id, lot.lot_name) for lot in material.lots if not lot.is_fraction]
    if form.validate_on_submit():
        lot_id = form.lot_id.data if form.lot_id.data != 0 else None
        
        # 予約後の予測重量を計算
        predicted_after_reserve = material.get_predicted_stock() - form.quantity.data
        
        # 最低重量を下回る場合は警告
        if predicted_after_reserve < material.min_weight:
            shortage = material.min_weight - predicted_after_reserve
            warning_msg = f'⚠️ 警告: この予約により予測在庫が最低量を{shortage:.1f}g下回ります（予測: {predicted_after_reserve:.1f}g / 最低: {material.min_weight:.1f}g）'
            
            # アクションタイプに応じた対処を促す
            if material.action_type == 'email' and material.email:
                warning_msg += f' → 購入担当者（{material.email}）にメール連絡してください'
            elif material.action_type == 'excel' and material.excel_path:
                warning_msg += f' → <a href="/open_excel/{material.id}" class="alert-link">発注用エクセルを開く</a>'
            
            flash(warning_msg, 'warning')
        
        reservation = Reservation(
            material_id=id, 
            lot_id=lot_id,
            lot_name=form.lot_name.data if form.lot_name.data else None,
            type='use', 
            quantity=form.quantity.data,
            user_name=form.user_name.data,
            purpose=form.purpose.data,
            scheduled_date=form.scheduled_date.data
        )
        db.session.add(reservation)
        db.session.commit()
        flash('使用予約を登録しました', 'success')
        return redirect(url_for('index'))
    contacts = Contact.query.order_by(Contact.name).all()
    return render_template('reserve.html', form=form, material=material, action='use', contacts=contacts)

@app.route('/reserve_replenish/<int:id>', methods=['GET', 'POST'])
def reserve_replenish(id):
    material = RawMaterial.query.get_or_404(id)
    form = ReservationForm()
    # ロット選択肢を追加（空欄も含む）
    form.lot_id.choices = [(0, '既存ロットから選択しない')] + [(lot.id, lot.lot_name) for lot in material.lots]
    if form.validate_on_submit():
        lot_id = form.lot_id.data if form.lot_id.data != 0 else None
        reservation = Reservation(
            material_id=id, 
            lot_id=lot_id,
            lot_name=form.lot_name.data if form.lot_name.data else None,
            type='replenish', 
            quantity=form.quantity.data,
            user_name=form.user_name.data,
            purpose=form.purpose.data,
            scheduled_date=form.scheduled_date.data
        )
        db.session.add(reservation)
        db.session.commit()
        flash('補充予約を登録しました', 'success')
        return redirect(url_for('index'))
    contacts = Contact.query.order_by(Contact.name).all()
    return render_template('reserve.html', form=form, material=material, action='replenish', contacts=contacts)

@app.route('/lots/<int:material_id>')
def lots(material_id):
    """原料のロット一覧"""
    material = RawMaterial.query.get_or_404(material_id)
    return render_template('lots.html', material=material)

@app.route('/add_lot/<int:material_id>', methods=['GET', 'POST'])
def add_lot(material_id):
    """ロット追加"""
    material = RawMaterial.query.get_or_404(material_id)
    form = LotForm()
    
    # 風袋の選択肢を設定
    tares = Tare.query.order_by(Tare.name).all()
    form.use_tare.choices = [(0, '風袋なし')] + [(t.id, f'{t.name} ({t.weight}g)') for t in tares]
    
    if form.validate_on_submit():
        # 風袋を使用する場合は、入力された重量から風袋重量を引く
        net_weight = form.weight.data
        if form.use_tare.data and form.use_tare.data != 0:
            tare = Tare.query.get(form.use_tare.data)
            if tare:
                net_weight = form.weight.data - tare.weight
        
        lot = Lot(material_id=material_id, lot_name=form.lot_name.data, weight=net_weight)
        db.session.add(lot)
        db.session.flush()  # lotのIDを取得するため
        
        # 統計用に実行済み予約を作成（ロット追加は補充扱い）
        if net_weight > 0:
            auto_reservation = Reservation(
                material_id=material_id,
                lot_id=lot.id,
                lot_name=form.lot_name.data,
                type='replenish',
                quantity=net_weight,
                actual_quantity=net_weight,
                user_name='システム',
                purpose=f'ロット追加（{form.lot_name.data}）',
                scheduled_date=datetime.now().date(),
                executed=True,
                executed_date=datetime.now()
            )
            db.session.add(auto_reservation)
        
        db.session.commit()
        flash(f'ロット「{form.lot_name.data}」を追加しました', 'success')
        return redirect(url_for('lots', material_id=material_id))
    return render_template('add_lot.html', form=form, material=material, tares=tares)

@app.route('/edit_lot/<int:id>', methods=['GET', 'POST'])
def edit_lot(id):
    """ロット編集"""
    lot = Lot.query.get_or_404(id)
    form = LotForm()
    
    # 風袋の選択肢を設定
    tares = Tare.query.order_by(Tare.name).all()
    form.use_tare.choices = [(0, '風袋なし')] + [(t.id, f'{t.name} ({t.weight}g)') for t in tares]
    
    if form.validate_on_submit():
        old_weight = lot.weight
        
        # 風袋を使用する場合は、入力された重量から風袋重量を引く
        new_weight = form.weight.data
        if form.use_tare.data and form.use_tare.data != 0:
            tare = Tare.query.get(form.use_tare.data)
            if tare:
                new_weight = form.weight.data - tare.weight
        
        lot.lot_name = form.lot_name.data
        lot.weight = new_weight
        
        # 重量が変化した場合、統計用に実行済み予約を作成
        weight_diff = new_weight - old_weight
        if weight_diff != 0:
            transaction_type = 'replenish' if weight_diff > 0 else 'use'
            quantity = abs(weight_diff)
            
            auto_reservation = Reservation(
                material_id=lot.material_id,
                lot_id=lot.id,
                lot_name=lot.lot_name,
                type=transaction_type,
                quantity=quantity,
                actual_quantity=quantity,
                user_name='システム',
                purpose=f'ロット直接編集（{lot.lot_name}）',
                scheduled_date=datetime.now().date(),
                executed=True,
                executed_date=datetime.now()
            )
            db.session.add(auto_reservation)
        
        db.session.commit()
        flash(f'ロット「{lot.lot_name}」を更新しました', 'success')
        return redirect(url_for('lots', material_id=lot.material_id))
    elif request.method == 'GET':
        form.lot_name.data = lot.lot_name
        form.weight.data = lot.weight
        form.use_tare.data = 0
    return render_template('edit_lot.html', form=form, lot=lot, tares=tares)

@app.route('/delete_lot/<int:id>')
def delete_lot(id):
    """ロット削除"""
    lot = Lot.query.get_or_404(id)
    material_id = lot.material_id
    lot_name = lot.lot_name
    lot_weight = lot.weight
    is_fraction = lot.is_fraction
    
    # このロットを参照しているReservationUsageレコードを削除
    ReservationUsage.query.filter_by(lot_id=id).delete()
    
    # 統計用に実行済み予約を作成（ロット削除は使用扱い）
    # ただし、端数処理されたロットは統計に含まれないため予約を作成しない
    if lot_weight > 0 and not is_fraction:
        auto_reservation = Reservation(
            material_id=material_id,
            lot_id=None,  # ロット削除後なのでNone
            lot_name=lot_name,
            type='use',
            quantity=lot_weight,
            actual_quantity=lot_weight,
            user_name='システム',
            purpose=f'ロット削除（{lot_name}）',
            scheduled_date=datetime.now().date(),
            executed=True,
            executed_date=datetime.now()
        )
        db.session.add(auto_reservation)
    
    db.session.delete(lot)
    db.session.commit()
    
    if is_fraction:
        flash(f'端数処理ロット「{lot_name}」を削除しました', 'success')
    else:
        flash(f'ロット「{lot_name}」を削除しました', 'success')
    
    return redirect(url_for('lots', material_id=material_id))

@app.route('/fraction_lot/<int:id>')
def fraction_lot(id):
    """ロットを端数処理"""
    lot = Lot.query.get_or_404(id)
    material_id = lot.material_id
    
    if lot.is_fraction:
        flash(f'ロット「{lot.lot_name}」は既に端数処理されています', 'warning')
    else:
        lot.is_fraction = True
        db.session.commit()
        flash(f'ロット「{lot.lot_name}」を端数処理しました', 'success')
    
    return redirect(url_for('lots', material_id=material_id))

@app.route('/unfraction_lot/<int:id>')
def unfraction_lot(id):
    """端数処理を解除して通常ロットに戻す"""
    lot = Lot.query.get_or_404(id)
    material_id = lot.material_id
    
    if not lot.is_fraction:
        flash(f'ロット「{lot.lot_name}」は通常のロットです', 'warning')
    else:
        lot.is_fraction = False
        db.session.commit()
        flash(f'ロット「{lot.lot_name}」を通常ロットに戻しました', 'success')
    
    return redirect(url_for('lots', material_id=material_id))

@app.route('/reservations')
def reservations():
    """予約管理ページ"""
    use_reservations = Reservation.query.filter_by(type='use', executed=False).order_by(Reservation.scheduled_date.asc(), Reservation.date.desc()).all()
    replenish_reservations = Reservation.query.filter_by(type='replenish', executed=False).order_by(Reservation.scheduled_date.asc(), Reservation.date.desc()).all()
    
    # 期限切れ予約を抽出
    overdue_reservations = [r for r in use_reservations + replenish_reservations if r.is_overdue()]
    
    # レシピ予約をグループ化
    recipe_groups = {}
    for r in use_reservations:
        if r.recipe_id:
            if r.recipe_id not in recipe_groups:
                recipe_groups[r.recipe_id] = {
                    'recipe': r.recipe,
                    'scheduled_date': r.scheduled_date,
                    'user_name': r.user_name,
                    'purpose': r.purpose,
                    'reservations': []
                }
            recipe_groups[r.recipe_id]['reservations'].append(r)
    
    return render_template('reservations.html', 
                         use_reservations=use_reservations,
                         replenish_reservations=replenish_reservations,
                         overdue_count=len(overdue_reservations),
                         recipe_groups=recipe_groups)

@app.route('/execute_reservation/<int:id>', methods=['GET', 'POST'])
def execute_reservation(id):
    """予約を実行して在庫に反映（複数ロット対応）"""
    reservation = Reservation.query.get_or_404(id)
    material = reservation.material
    
    # POSTリクエストの場合、実際の量とロット情報を取得
    if request.method == 'POST':
        try:
            if reservation.type == 'use':
                # 複数ロットの情報を取得
                lot_usages = []
                index = 0
                while True:
                    lot_id_key = f'lot_id_{index}'
                    lot_quantity_key = f'lot_quantity_{index}'
                    
                    lot_id = request.form.get(lot_id_key)
                    lot_quantity = request.form.get(lot_quantity_key)
                    
                    if not lot_id or not lot_quantity:
                        break
                    
                    lot_usages.append({
                        'lot_id': int(lot_id),
                        'quantity': float(lot_quantity)
                    })
                    index += 1
                
                if not lot_usages:
                    flash('エラー: ロットを選択してください', 'danger')
                    return redirect(url_for('reservations'))
                
                # 合計使用量を計算
                total_quantity = sum(usage['quantity'] for usage in lot_usages)
                reservation.actual_quantity = total_quantity
                
                # 各ロットから在庫を減少
                for usage in lot_usages:
                    lot = Lot.query.get(usage['lot_id'])
                    if not lot:
                        flash(f'エラー: ロットが見つかりません', 'danger')
                        db.session.rollback()
                        return redirect(url_for('reservations'))
                    
                    if lot.weight >= usage['quantity']:
                        lot.weight -= usage['quantity']
                    else:
                        flash(f'エラー: ロット「{lot.lot_name}」の在庫が不足しています', 'danger')
                        db.session.rollback()
                        return redirect(url_for('reservations'))
                    
                    # ReservationUsageレコードを作成
                    reservation_usage = ReservationUsage(
                        reservation_id=reservation.id,
                        lot_id=usage['lot_id'],
                        quantity=usage['quantity']
                    )
                    db.session.add(reservation_usage)
            
            elif reservation.type == 'replenish':
                # 補充予約の場合
                actual_quantity = float(request.form.get('actual_quantity', reservation.quantity))
                reservation.actual_quantity = actual_quantity
                quantity_to_use = actual_quantity
                
                # ロット名を取得
                lot_name = request.form.get('lot_name', '').strip()
                if not lot_name:
                    flash('エラー: ロット名を入力してください', 'danger')
                    return redirect(url_for('reservations'))
                
                reservation.lot_name = lot_name
                
                # 新規ロット名が指定されている場合
                existing_lot = Lot.query.filter_by(material_id=material.id, lot_name=lot_name).first()
                if existing_lot:
                    # 既存ロットに追加
                    existing_lot.weight += quantity_to_use
                else:
                    # 新規ロット作成
                    new_lot = Lot(material_id=material.id, lot_name=lot_name, weight=quantity_to_use)
                    db.session.add(new_lot)
            
            # 予約を実行済みにマーク
            reservation.executed = True
            reservation.executed_date = datetime.now()
            db.session.commit()
            flash(f'予約を実行しました: {material.name}', 'success')
        
        except Exception as e:
            db.session.rollback()
            flash(f'エラーが発生しました: {str(e)}', 'danger')
        
        return redirect(url_for('reservations'))
    
    # GETリクエストの場合、実行フォームを表示
    return render_template('execute_reservation.html', reservation=reservation)

@app.route('/execute_recipe/<int:recipe_id>', methods=['POST'])
def execute_recipe(recipe_id):
    """レシピ予約を一括実行（複数ロット対応）"""
    recipe = Recipe.query.get_or_404(recipe_id)
    
    # このレシピに紐づく未実行の使用予約を取得
    reservations = Reservation.query.filter_by(
        recipe_id=recipe_id,
        type='use',
        executed=False
    ).all()
    
    if not reservations:
        flash('実行する予約が見つかりません', 'warning')
        return redirect(url_for('reservations'))
    
    try:
        # 各原料の実績値を取得して実行
        for reservation in reservations:
            material = reservation.material
            
            # 複数ロットの情報を取得（lot_id_<reservation_id>_<index> と lot_quantity_<reservation_id>_<index>）
            lot_usages = []
            index = 0
            while True:
                lot_id_key = f'lot_id_{reservation.id}_{index}'
                lot_quantity_key = f'lot_quantity_{reservation.id}_{index}'
                
                lot_id = request.form.get(lot_id_key)
                lot_quantity = request.form.get(lot_quantity_key)
                
                if not lot_id or not lot_quantity:
                    break
                
                lot_usages.append({
                    'lot_id': int(lot_id),
                    'quantity': float(lot_quantity)
                })
                index += 1
            
            if not lot_usages:
                flash(f'エラー: {material.name}のロットを選択してください', 'danger')
                db.session.rollback()
                return redirect(url_for('reservations'))
            
            # 合計使用量を計算
            total_quantity = sum(usage['quantity'] for usage in lot_usages)
            reservation.actual_quantity = total_quantity
            
            # 各ロットから在庫を減少
            for usage in lot_usages:
                lot = Lot.query.get(usage['lot_id'])
                if not lot:
                    flash(f'エラー: ロットが見つかりません', 'danger')
                    db.session.rollback()
                    return redirect(url_for('reservations'))
                
                if lot.weight >= usage['quantity']:
                    lot.weight -= usage['quantity']
                else:
                    flash(f'エラー: ロット「{lot.lot_name}」の在庫が不足しています', 'danger')
                    db.session.rollback()
                    return redirect(url_for('reservations'))
                
                # ReservationUsageレコードを作成
                reservation_usage = ReservationUsage(
                    reservation_id=reservation.id,
                    lot_id=usage['lot_id'],
                    quantity=usage['quantity']
                )
                db.session.add(reservation_usage)
            
            # 予約を実行済みにマーク
            reservation.executed = True
            reservation.executed_date = datetime.now()
        
        db.session.commit()
        flash(f'レシピ「{recipe.name}」の予約を一括実行しました', 'success')
    
    except Exception as e:
        db.session.rollback()
        flash(f'エラーが発生しました: {str(e)}', 'danger')
    
    return redirect(url_for('reservations'))

@app.route('/edit_reservation/<int:id>', methods=['GET', 'POST'])
def edit_reservation(id):
    """予約編集"""
    reservation = Reservation.query.get_or_404(id)
    
    if request.method == 'POST':
        reservation.user_name = request.form.get('user_name', '')
        reservation.purpose = request.form.get('purpose', '')
        reservation.quantity = float(request.form.get('quantity', reservation.quantity))
        
        # 予定日の処理
        scheduled_date_str = request.form.get('scheduled_date')
        if scheduled_date_str:
            reservation.scheduled_date = datetime.strptime(scheduled_date_str, '%Y-%m-%d').date()
        
        # ロット名/ロット選択の処理
        if reservation.type == 'use':
            lot_id = request.form.get('lot_id')
            if lot_id:
                reservation.lot_id = int(lot_id)
        else:  # replenish
            lot_name = request.form.get('lot_name', '').strip()
            if lot_name:
                reservation.lot_name = lot_name
        
        db.session.commit()
        flash('予約を更新しました', 'success')
        return redirect(url_for('reservations'))
    
    contacts = Contact.query.order_by(Contact.name).all()
    return render_template('edit_reservation.html', reservation=reservation, contacts=contacts)

@app.route('/delete_reservation/<int:id>')
def delete_reservation(id):
    """予約削除"""
    reservation = Reservation.query.get_or_404(id)
    db.session.delete(reservation)
    db.session.commit()
    flash('予約を削除しました', 'success')
    return redirect(url_for('reservations'))

@app.route('/export')
def export():
    materials = RawMaterial.query.all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', '名前', '現在重量', '単位', '最低量', '予測在庫'])
    for material in materials:
        current_weight = material.get_total_lot_weight()
        predicted = material.get_predicted_stock()
        writer.writerow([material.id, material.name, round(current_weight, 2), material.unit, material.min_weight, round(predicted, 2)])
    output.seek(0)
    # BOM付きUTF-8でエンコードして日本語文字化けを防止
    csv_data = '\ufeff' + output.getvalue()
    response = make_response(csv_data)
    response.headers['Content-Disposition'] = 'attachment; filename=inventory.csv'
    response.headers['Content-type'] = 'text/csv; charset=utf-8-sig'
    return response

@app.route('/send_alert_email/<int:id>', methods=['POST'])
def send_alert_email(id):
    """アラートメールを送信"""
    material = RawMaterial.query.get_or_404(id)
    
    if not material.email:
        flash('購入担当者のメールアドレスが登録されていません。', 'warning')
        return redirect(url_for('index'))
    
    try:
        # メール内容
        current_weight = material.get_total_lot_weight()
        predicted_stock = material.get_predicted_stock()
        subject = f"【在庫アラート】{material.name}の補充が必要です"
        body = f"""
在庫管理システムからの自動通知

原料名: {material.name}
現在量: {current_weight:.2f} {material.unit}
最低量: {material.min_weight} {material.unit}
予測在庫量: {predicted_stock:.2f} {material.unit}

予測在庫量が最低量を下回る見込みです。
至急、補充の手配をお願いします。

※このメールは在庫管理システムから自動送信されています。
        """
        
        # 実際のメール送信（Gmail使用例）
        # 注意: 本番環境では環境変数やconfigファイルで設定してください
        sender_email = "your-email@gmail.com"  # 送信元メール
        sender_password = "your-app-password"  # アプリパスワード
        
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = material.email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # Gmail SMTPサーバー経由で送信（実際の送信を有効にする場合はコメント解除）
        # server = smtplib.SMTP('smtp.gmail.com', 587)
        # server.starttls()
        # server.login(sender_email, sender_password)
        # server.send_message(msg)
        # server.quit()
        
        # デモ用: 実際には送信せずにメッセージのみ表示
        flash(f'アラートメールを {material.email} に送信しました（デモモード）', 'success')
        
    except Exception as e:
        flash(f'メール送信に失敗しました: {str(e)}', 'danger')
    
    return redirect(url_for('index'))

# ======================================================
# ラベル管理ルート (MaterialLabel CRUD)
# ======================================================

@app.route('/labels')
def labels():
    all_labels = MaterialLabel.query.order_by(MaterialLabel.name).all()
    return render_template('labels.html', labels=all_labels)

@app.route('/labels/add', methods=['POST'])
def add_label():
    name  = request.form.get('name', '').strip()
    color = request.form.get('color', '#6c757d').strip()
    if not name:
        flash('ラベル名を入力してください', 'warning')
        return redirect(url_for('labels'))
    if MaterialLabel.query.filter_by(name=name).first():
        flash(f'ラベル名「{name}」は既に使用されています', 'warning')
        return redirect(url_for('labels'))
    db.session.add(MaterialLabel(name=name, color=color))
    db.session.commit()
    flash(f'ラベル「{name}」を追加しました', 'success')
    return redirect(url_for('labels'))

@app.route('/labels/edit/<int:id>', methods=['GET', 'POST'])
def edit_label(id):
    label = MaterialLabel.query.get_or_404(id)
    if request.method == 'POST':
        name  = request.form.get('name', '').strip()
        color = request.form.get('color', label.color).strip()
        if not name:
            flash('ラベル名を入力してください', 'warning')
            return redirect(url_for('edit_label', id=id))
        conflict = MaterialLabel.query.filter(
            MaterialLabel.name == name, MaterialLabel.id != id
        ).first()
        if conflict:
            flash(f'ラベル名「{name}」は既に使用されています', 'warning')
            return redirect(url_for('edit_label', id=id))
        label.name  = name
        label.color = color
        db.session.commit()
        flash(f'ラベル「{name}」を更新しました', 'success')
        return redirect(url_for('labels'))
    return render_template('edit_label.html', label=label)

@app.route('/labels/delete/<int:id>', methods=['POST'])
def delete_label(id):
    label = MaterialLabel.query.get_or_404(id)
    name  = label.name
    RawMaterial.query.filter_by(label_id=id).update({'label_id': None})
    db.session.delete(label)
    db.session.commit()
    flash(f'ラベル「{name}」を削除しました', 'success')
    return redirect(url_for('labels'))

# ======================================================
# 物性値テンプレート管理ルート
# ======================================================

@app.route('/property_templates')
def property_templates():
    """物性値テンプレート一覧: ラベルごとにフィールド一覧を表示"""
    labels = MaterialLabel.query.order_by(MaterialLabel.name).all()
    unlabeled_fields = PropertyField.query.filter_by(label_id=None).order_by(PropertyField.order_index).all()
    return render_template('property_templates.html', labels=labels,
                           unlabeled_fields=unlabeled_fields)

@app.route('/property_templates/fields/add', methods=['POST'])
def add_property_field():
    """フィールドを追加"""
    label_id_raw = request.form.get('label_id', '')
    label_id     = int(label_id_raw) if label_id_raw else None
    name         = request.form.get('name', '').strip()
    field_type   = request.form.get('field_type', 'string')
    unit         = request.form.get('unit', '').strip() or None
    if not name:
        flash('フィールド名を入力してください', 'warning')
        return redirect(url_for('property_templates'))
    max_order = db.session.query(db.func.max(PropertyField.order_index)).filter_by(label_id=label_id).scalar() or 0
    field = PropertyField(label_id=label_id, name=name, field_type=field_type,
                          unit=unit, order_index=max_order + 1)
    db.session.add(field)
    db.session.commit()
    flash(f'フィールド「{name}」を追加しました', 'success')
    return redirect(url_for('property_templates'))

@app.route('/property_templates/fields/edit/<int:id>', methods=['GET', 'POST'])
def edit_property_field(id):
    """フィールドを編集"""
    field = PropertyField.query.get_or_404(id)
    if request.method == 'POST':
        name       = request.form.get('name', '').strip()
        field_type = request.form.get('field_type', field.field_type)
        unit       = request.form.get('unit', '').strip() or None
        if not name:
            flash('フィールド名を入力してください', 'warning')
            return redirect(url_for('edit_property_field', id=id))
        field.name       = name
        field.field_type = field_type
        field.unit       = unit
        db.session.commit()
        flash(f'フィールド「{name}」を更新しました', 'success')
        return redirect(url_for('property_templates'))
    return render_template('edit_property_field.html', field=field)

@app.route('/property_templates/fields/delete/<int:id>', methods=['POST'])
def delete_property_field(id):
    """フィールドを削除（関連する物性値も連鎖削除）"""
    field = PropertyField.query.get_or_404(id)
    name  = field.name
    db.session.delete(field)
    db.session.commit()
    flash(f'フィールド「{name}」を削除しました', 'success')
    return redirect(url_for('property_templates'))

@app.route('/property_templates/fields/reorder', methods=['POST'])
def reorder_property_fields():
    """フィールド並び順をJSONで更新（AJAX用）"""
    data = request.get_json()
    for item in data:
        field = PropertyField.query.get(item['id'])
        if field:
            field.order_index = item['order']
    db.session.commit()
    return jsonify({'status': 'ok'})

# ======================================================
# ロット物性値ルート
# ======================================================

@app.route('/lots/<int:lot_id>/properties', methods=['GET', 'POST'])
def lot_properties(lot_id):
    """ロットの物性値を表示・編集"""
    lot      = Lot.query.get_or_404(lot_id)
    material = lot.material
    label_id = material.label_id
    fields   = PropertyField.query.filter_by(label_id=label_id).order_by(PropertyField.order_index).all()

    if request.method == 'POST':
        for field in fields:
            raw = request.form.get(f'field_{field.id}', '').strip()
            prop = LotProperty.query.filter_by(lot_id=lot_id, field_id=field.id).first()
            if prop is None:
                prop = LotProperty(lot_id=lot_id, field_id=field.id)
                db.session.add(prop)
            if field.field_type == 'number':
                try:
                    prop.value_number = float(raw) if raw else None
                    prop.value_string = None
                except ValueError:
                    prop.value_number = None
                    prop.value_string = raw
            elif field.field_type == 'date':
                prop.value_string = raw or None   # 'YYYY-MM-DD' 形式で保存
                prop.value_number = None
            else:
                prop.value_string = raw or None
                prop.value_number = None
            prop.date_updated = datetime.utcnow()
        db.session.commit()
        flash('物性値を保存しました', 'success')
        return redirect(url_for('lot_properties', lot_id=lot_id))

    existing = {lp.field_id: lp for lp in lot.properties}
    return render_template('lot_properties.html', lot=lot, material=material,
                           fields=fields, existing=existing)

# ======================================================
# Excelエクスポートルート
# ======================================================

@app.route('/export/properties')
def export_properties():
    """全原料のロット物性値をラベル別シートでExcelエクスポート"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = Workbook()
    wb.remove(wb.active)

    def make_sheet(wb, sheet_name, label_id, fields):
        ws = wb.create_sheet(title=sheet_name[:31])
        headers = ['原料名', 'ロット名', '重量(g)', '端数'] + [
            f.name + (f' ({f.unit})' if f.unit else '') for f in fields
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='2563EB')
            cell.alignment = Alignment(horizontal='center')
        if label_id is None:
            mat_list = RawMaterial.query.filter_by(label_id=None).order_by(RawMaterial.name).all()
        else:
            mat_list = RawMaterial.query.filter_by(label_id=label_id).order_by(RawMaterial.name).all()
        for material in mat_list:
            for lot in sorted(material.lots, key=lambda l: l.lot_name):
                row = [material.name, lot.lot_name, lot.weight,
                       '端数' if lot.is_fraction else '']
                prop_map = {lp.field_id: lp for lp in lot.properties}
                for field in fields:
                    lp = prop_map.get(field.id)
                    if lp:
                        if field.field_type == 'number':
                            row.append(lp.value_number)
                        elif field.field_type == 'date' and lp.value_string:
                            # 日付型: Excel日付セルとして出力
                            try:
                                from datetime import date as _date
                                row.append(_date.fromisoformat(lp.value_string))
                            except Exception:
                                row.append(lp.value_string)
                        else:
                            row.append(lp.value_string or '')
                    else:
                        row.append('')
                ws.append(row)
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(12, min(max_len + 4, 50))
        return ws

    all_labels = MaterialLabel.query.order_by(MaterialLabel.name).all()
    for label in all_labels:
        fields = PropertyField.query.filter_by(label_id=label.id).order_by(PropertyField.order_index).all()
        make_sheet(wb, label.name, label.id, fields)

    unlabeled_fields = PropertyField.query.filter_by(label_id=None).order_by(PropertyField.order_index).all()
    make_sheet(wb, 'ラベルなし', None, unlabeled_fields)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='物性値データベース.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ======================================================

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/api/stats')
def api_stats():
    # ラベルフィルタ対応
    label_id_param = request.args.get('label_id', '')
    mat_query = RawMaterial.query
    if label_id_param == 'none':
        mat_query = mat_query.filter(RawMaterial.label_id == None)
    elif label_id_param:
        try:
            mat_query = mat_query.filter(RawMaterial.label_id == int(label_id_param))
        except ValueError:
            pass
    materials = mat_query.all()

    # 総在庫数
    total_materials = len(materials)

    # 低在庫アラート数（予測在庫で判定）
    low_stock_count = sum(1 for m in materials if m.is_low_stock_alert())

    # アラート一覧
    alert_materials = []
    for material in materials:
        if material.is_low_stock_alert():
            total_weight = material.get_total_lot_weight()
            predicted = material.get_predicted_stock()
            critical_periods = material.get_critical_periods()

            # 日付をJSON互換形式に変換
            serialized_periods = []
            for period in critical_periods:
                serialized_periods.append({
                    'start_date': period['start_date'].isoformat() if period['start_date'] else None,
                    'end_date': period['end_date'].isoformat() if period['end_date'] else None,
                    'min_stock': round(period['min_stock'], 2),
                    'shortage': round(period['shortage'], 2)
                })

            alert_materials.append({
                'id': material.id,
                'name': material.name,
                'current': round(total_weight, 2),
                'predicted': round(predicted, 2),
                'min_weight': material.min_weight,
                'unit': material.unit,
                'email': material.email,
                'excel_path': material.excel_path,
                'action_type': material.action_type,
                'critical_periods': serialized_periods,
                'label_name':  material.label.name  if material.label else None,
                'label_color': material.label.color if material.label else None
            })

    # 在庫状況データ
    materials_data = []
    for material in materials:
        total_weight = material.get_total_lot_weight()
        predicted_stock = material.get_predicted_stock()
        materials_data.append({
            'name': material.name,
            'current': round(total_weight, 2),
            'predicted': round(predicted_stock, 2),
            'min_weight': material.min_weight,
            'unit': material.unit,
            'label_name':  material.label.name  if material.label else None,
            'label_color': material.label.color if material.label else None
        })

    # 予約情報の集計
    use_reservations = Reservation.query.filter_by(type='use').order_by(Reservation.date.desc()).limit(5).all()
    replenish_reservations = Reservation.query.filter_by(type='replenish').order_by(Reservation.date.desc()).limit(5).all()

    use_list = [{
        'material': r.material.name,
        'lot': r.lot.lot_name if r.lot else '原料全体',
        'quantity': r.quantity,
        'date': r.date.strftime('%Y/%m/%d %H:%M') if r.date else 'N/A'
    } for r in use_reservations]

    replenish_list = [{
        'material': r.material.name,
        'lot': r.lot.lot_name if r.lot else '原料全体',
        'quantity': r.quantity,
        'date': r.date.strftime('%Y/%m/%d %H:%M') if r.date else 'N/A'
    } for r in replenish_reservations]

    # 期限切れ予約数を計算
    from datetime import date, timedelta
    today = date.today()
    all_reservations = Reservation.query.filter_by(executed=False).all()
    overdue_count = sum(1 for r in all_reservations if r.scheduled_date and r.scheduled_date < today)

    # 今週（7日以内）の予約数を計算
    week_later = today + timedelta(days=7)
    week_reservations = sum(1 for r in all_reservations if r.scheduled_date and today <= r.scheduled_date <= week_later)

    # 全ラベル一覧（ダッシュボードのドロップダウン用）
    all_labels = [{'id': l.id, 'name': l.name, 'color': l.color}
                  for l in MaterialLabel.query.order_by(MaterialLabel.name).all()]

    return jsonify({
        'total_materials': total_materials,
        'low_stock_count': low_stock_count,
        'alert_materials': alert_materials,
        'materials': materials_data,
        'use_reservations': use_list,
        'replenish_reservations': replenish_list,
        'overdue_count': overdue_count,
        'week_reservations': week_reservations,
        'labels': all_labels
    })

@app.route('/api/material_stats/<int:id>')
def api_material_stats(id):
    """原料の期間別統計データを取得"""
    material = RawMaterial.query.get_or_404(id)
    
    # 各期間の統計を取得
    stats = {
        '1d': material.get_usage_stats(1),
        '7d': material.get_usage_stats(7),
        '1m': material.get_usage_stats(30),
        '3m': material.get_usage_stats(90),
        '6m': material.get_usage_stats(180),
        '1y': material.get_usage_stats(365)
    }
    
    return jsonify({
        'material_id': material.id,
        'material_name': material.name,
        'current_stock': material.get_total_lot_weight(),
        'unit': material.unit,
        'stats': stats
    })

# Recipe Management Routes
@app.route('/recipes')
def recipes():
    recipes = Recipe.query.order_by(Recipe.date_created.desc()).all()
    return render_template('recipes.html', recipes=recipes)

@app.route('/add_recipe', methods=['GET', 'POST'])
def add_recipe():
    form = RecipeForm()
    materials = RawMaterial.query.all()
    if form.validate_on_submit():
        recipe = Recipe(
            name=form.name.data,
            description=form.description.data,
            type='use'
        )
        db.session.add(recipe)
        db.session.flush()  # Get recipe.id before adding items
        
        # Add recipe items from form data
        for material in materials:
            quantity_key = f'material_{material.id}_quantity'
            lot_name_key = f'material_{material.id}_lot_name'
            quantity = request.form.get(quantity_key, type=float)
            lot_name = request.form.get(lot_name_key, '')
            
            if quantity and quantity > 0:
                recipe_item = RecipeItem(
                    recipe_id=recipe.id,
                    material_id=material.id,
                    quantity=quantity,
                    lot_name=lot_name if lot_name else None
                )
                db.session.add(recipe_item)
        
        db.session.commit()
        flash('レシピを登録しました', 'success')
        return redirect(url_for('recipes'))
    
    return render_template('add_recipe.html', form=form, materials=materials)

@app.route('/edit_recipe/<int:id>', methods=['GET', 'POST'])
def edit_recipe(id):
    recipe = Recipe.query.get_or_404(id)
    form = RecipeForm()
    materials = RawMaterial.query.all()
    
    if form.validate_on_submit():
        recipe.name = form.name.data
        recipe.description = form.description.data
        
        # Delete existing recipe items
        RecipeItem.query.filter_by(recipe_id=recipe.id).delete()
        
        # Add new recipe items
        for material in materials:
            quantity_key = f'material_{material.id}_quantity'
            lot_name_key = f'material_{material.id}_lot_name'
            quantity = request.form.get(quantity_key, type=float)
            lot_name = request.form.get(lot_name_key, '')
            
            if quantity and quantity > 0:
                recipe_item = RecipeItem(
                    recipe_id=recipe.id,
                    material_id=material.id,
                    quantity=quantity,
                    lot_name=lot_name if lot_name else None
                )
                db.session.add(recipe_item)
        
        db.session.commit()
        flash('レシピを更新しました', 'success')
        return redirect(url_for('recipes'))
    
    if request.method == 'GET':
        form.name.data = recipe.name
        form.description.data = recipe.description
    
    return render_template('edit_recipe.html', form=form, recipe=recipe, materials=materials)

@app.route('/delete_recipe/<int:id>')
def delete_recipe(id):
    recipe = Recipe.query.get_or_404(id)
    RecipeItem.query.filter_by(recipe_id=recipe.id).delete()
    db.session.delete(recipe)
    db.session.commit()
    flash('レシピを削除しました', 'success')
    return redirect(url_for('recipes'))

@app.route('/use_recipe/<int:recipe_id>', methods=['POST'])
@csrf.exempt
def use_recipe(recipe_id):
    recipe = Recipe.query.get_or_404(recipe_id)
    user_name = request.form.get('user_name', '')
    purpose = request.form.get('purpose', '')
    scheduled_date_str = request.form.get('scheduled_date', '')
    scheduled_date = datetime.strptime(scheduled_date_str, '%Y-%m-%d').date() if scheduled_date_str else None
    
    # Create reservations for each item in the recipe
    for item in recipe.items:
        reservation = Reservation(
            material_id=item.material_id,
            recipe_id=recipe.id,
            lot_name=item.lot_name,
            type='use',
            quantity=item.quantity,
            user_name=user_name,
            purpose=purpose,
            scheduled_date=scheduled_date
        )
        db.session.add(reservation)
    
    db.session.commit()
    flash(f'レシピ「{recipe.name}」から使用予約を作成しました', 'success')
    return redirect(url_for('reservations'))

# Backup Management Routes
def get_backup_folder():
    """バックアップフォルダのパスを取得"""
    config = load_config()
    if 'database_folder' in config:
        return os.path.join(config['database_folder'], 'backups')
    return 'backups'

def get_db_path():
    """データベースファイルのパスを取得"""
    return app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')

def ensure_backup_folder():
    """バックアップフォルダの存在を確認し、なければ作成"""
    backup_folder = get_backup_folder()
    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)

@app.route('/backup')
def backup_management():
    """バックアップ管理ページ"""
    ensure_backup_folder()
    backups = []
    backup_folder = get_backup_folder()
    
    if os.path.exists(backup_folder):
        for filename in os.listdir(backup_folder):
            if filename.endswith('.db'):
                filepath = os.path.join(backup_folder, filename)
                stat = os.stat(filepath)
                backups.append({
                    'filename': filename,
                    'size': stat.st_size / 1024,  # KB
                    'created': datetime.fromtimestamp(stat.st_mtime).strftime('%Y/%m/%d %H:%M:%S')
                })
    
    backups.sort(key=lambda x: x['created'], reverse=True)
    return render_template('backup.html', backups=backups)

@app.route('/backup/create', methods=['POST'])
def create_backup():
    """新規バックアップを作成"""
    try:
        ensure_backup_folder()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'inventory_backup_{timestamp}.db'
        backup_folder = get_backup_folder()
        backup_path = os.path.join(backup_folder, backup_filename)
        db_path = get_db_path()
        
        if os.path.exists(db_path):
            shutil.copy2(db_path, backup_path)
            flash(f'バックアップを作成しました: {backup_filename}', 'success')
        else:
            flash('データベースファイルが見つかりません', 'danger')
    except Exception as e:
        flash(f'バックアップの作成に失敗しました: {str(e)}', 'danger')
    
    return redirect(url_for('backup_management'))

@app.route('/backup/restore/<filename>', methods=['POST'])
def restore_backup(filename):
    """バックアップから復元"""
    try:
        backup_folder = get_backup_folder()
        backup_path = os.path.join(backup_folder, filename)
        db_path = get_db_path()
        
        if not os.path.exists(backup_path):
            flash('指定されたバックアップファイルが見つかりません', 'danger')
            return redirect(url_for('backup_management'))
        
        # 現在のDBをバックアップ（復元前の安全策）
        if os.path.exists(db_path):
            db_folder = os.path.dirname(db_path)
            safety_backup = os.path.join(db_folder, f'inventory_before_restore_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db')
            shutil.copy2(db_path, safety_backup)
        
        # バックアップから復元
        shutil.copy2(backup_path, db_path)
        flash(f'バックアップから復元しました: {filename}', 'success')
    except Exception as e:
        flash(f'復元に失敗しました: {str(e)}', 'danger')
    
    return redirect(url_for('backup_management'))

@app.route('/backup/download/<filename>')
def download_backup(filename):
    """バックアップファイルをダウンロード"""
    try:
        backup_folder = get_backup_folder()
        backup_path = os.path.join(backup_folder, filename)
        if os.path.exists(backup_path):
            return send_file(backup_path, as_attachment=True, download_name=filename)
        else:
            flash('指定されたバックアップファイルが見つかりません', 'danger')
            return redirect(url_for('backup_management'))
    except Exception as e:
        flash(f'ダウンロードに失敗しました: {str(e)}', 'danger')
        return redirect(url_for('backup_management'))

@app.route('/backup/delete/<filename>', methods=['POST'])
def delete_backup(filename):
    """バックアップファイルを削除"""
    try:
        backup_folder = get_backup_folder()
        backup_path = os.path.join(backup_folder, filename)
        if os.path.exists(backup_path):
            os.remove(backup_path)
            flash(f'バックアップを削除しました: {filename}', 'success')
        else:
            flash('指定されたバックアップファイルが見つかりません', 'danger')
    except Exception as e:
        flash(f'削除に失敗しました: {str(e)}', 'danger')
    
    return redirect(url_for('backup_management'))

@app.route('/open_excel/<int:id>')
def open_excel(id):
    """指定された原料のエクセルファイルを開く"""
    material = RawMaterial.query.get_or_404(id)
    
    if not material.excel_path:
        flash('エクセルファイルパスが登録されていません', 'warning')
        return redirect(url_for('dashboard'))
    
    try:
        # ファイルの存在確認
        if not os.path.exists(material.excel_path):
            flash(f'ファイルが見つかりません: {material.excel_path}', 'danger')
            return redirect(url_for('dashboard'))
        
        # OSに応じてファイルを開く
        if platform.system() == 'Windows':
            os.startfile(material.excel_path)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.call(['open', material.excel_path])
        else:  # Linux
            subprocess.call(['xdg-open', material.excel_path])
        
        flash(f'{material.name}のエクセルファイルを開きました', 'success')
    except Exception as e:
        flash(f'ファイルを開けませんでした: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/settings')
def settings():
    """設定画面"""
    config = load_config()
    current_db_folder = config.get('database_folder', 'デフォルト（instance）')
    current_db_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
    
    return render_template('settings.html', 
                         db_folder=current_db_folder,
                         db_path=current_db_path)

@app.route('/tare_settings')
def tare_settings():
    """風袋設定画面"""
    tares = Tare.query.order_by(Tare.name).all()
    return render_template('tare_settings.html', tares=tares)

@app.route('/add_tare', methods=['GET', 'POST'])
def add_tare():
    """風袋を追加"""
    form = TareForm()
    if form.validate_on_submit():
        tare = Tare(
            name=form.name.data,
            weight=form.weight.data,
            description=form.description.data
        )
        db.session.add(tare)
        db.session.commit()
        flash(f'風袋「{form.name.data}」を追加しました', 'success')
        return redirect(url_for('tare_settings'))
    return render_template('add_tare.html', form=form)

@app.route('/edit_tare/<int:id>', methods=['GET', 'POST'])
def edit_tare(id):
    """風袋を編集"""
    tare = Tare.query.get_or_404(id)
    form = TareForm()
    if form.validate_on_submit():
        tare.name = form.name.data
        tare.weight = form.weight.data
        tare.description = form.description.data
        db.session.commit()
        flash(f'風袋「{tare.name}」を更新しました', 'success')
        return redirect(url_for('tare_settings'))
    elif request.method == 'GET':
        form.name.data = tare.name
        form.weight.data = tare.weight
        form.description.data = tare.description
    return render_template('edit_tare.html', form=form, tare=tare)

@app.route('/delete_tare/<int:id>')
def delete_tare(id):
    """風袋を削除"""
    tare = Tare.query.get_or_404(id)
    tare_name = tare.name
    db.session.delete(tare)
    db.session.commit()
    flash(f'風袋「{tare_name}」を削除しました', 'success')
    return redirect(url_for('tare_settings'))

@app.route('/contacts')
def contacts():
    """アドレス帳一覧"""
    contacts = Contact.query.order_by(Contact.name).all()
    return render_template('contacts.html', contacts=contacts)

@app.route('/add_contact', methods=['GET', 'POST'])
def add_contact():
    """連絡先を追加"""
    form = ContactForm()
    if form.validate_on_submit():
        contact = Contact(
            name=form.name.data,
            email=form.email.data,
            description=form.description.data
        )
        db.session.add(contact)
        db.session.commit()
        flash(f'連絡先「{form.name.data}」を追加しました', 'success')
        return redirect(url_for('contacts'))
    return render_template('add_contact.html', form=form)

@app.route('/edit_contact/<int:id>', methods=['GET', 'POST'])
def edit_contact(id):
    """連絡先を編集"""
    contact = Contact.query.get_or_404(id)
    form = ContactForm()
    if form.validate_on_submit():
        contact.name = form.name.data
        contact.email = form.email.data
        contact.description = form.description.data
        db.session.commit()
        flash(f'連絡先「{contact.name}」を更新しました', 'success')
        return redirect(url_for('contacts'))
    elif request.method == 'GET':
        form.name.data = contact.name
        form.email.data = contact.email
        form.description.data = contact.description
    return render_template('edit_contact.html', form=form, contact=contact)

@app.route('/delete_contact/<int:id>')
def delete_contact(id):
    """連絡先を削除"""
    contact = Contact.query.get_or_404(id)
    contact_name = contact.name
    db.session.delete(contact)
    db.session.commit()
    flash(f'連絡先「{contact_name}」を削除しました', 'success')
    return redirect(url_for('contacts'))

@app.route('/change_database_folder', methods=['POST'])
def change_database_folder():
    """データベースフォルダを変更"""
    try:
        from tkinter import Tk, filedialog, messagebox
        
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        folder = filedialog.askdirectory(
            title='新しいデータベースフォルダを選択してください',
            initialdir=os.path.dirname(app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', ''))
        )
        
        root.destroy()
        
        if folder:
            config = load_config()
            config['database_folder'] = folder
            save_config(config)
            
            flash(f'データベースフォルダを変更しました: {folder}\nアプリを再起動してください。', 'success')
        else:
            flash('フォルダが選択されませんでした', 'warning')
    except Exception as e:
        flash(f'エラー: {str(e)}', 'danger')
    
    return redirect(url_for('settings'))

# ======================================================
# 作業工程進捗管理ルート
# ======================================================

@app.route('/work_orders')
def work_orders():
    """作業工程一覧（ページロード時に全作業のステータスを自動同期）"""
    # 非アーカイブ作業のステータスを実際の日付から自動同期
    all_active = WorkOrder.query.filter_by(archived=False).all()
    changed = False
    for wo in all_active:
        new_status = wo.compute_status()
        if wo.status != new_status:
            wo.status = new_status
            changed = True
    if changed:
        db.session.commit()

    show_archived = request.args.get('archived', '0') == '1'
    status_filter = request.args.get('status', 'all')
    priority_filter = request.args.get('priority', 'all')
    type_filter = request.args.get('type', 'all')
    search = request.args.get('search', '')

    query = WorkOrder.query.filter_by(archived=show_archived)

    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    if priority_filter != 'all':
        query = query.filter_by(priority=priority_filter)
    if type_filter != 'all':
        query = query.filter_by(experiment_type=type_filter)
    if search:
        query = query.filter(
            db.or_(
                WorkOrder.process_name.contains(search),
                WorkOrder.lot_name.contains(search)
            )
        )

    # 優先度→完了予定日の順でソート
    orders = query.order_by(
        db.case(
            {'critical': 0, 'high': 1, 'low': 2, 'none': 3},
            value=WorkOrder.priority
        ),
        WorkOrder.planned_end.asc().nullslast(),
        WorkOrder.date_created.desc()
    ).all()

    return render_template('work_orders.html',
                           work_orders=orders,
                           show_archived=show_archived,
                           status_filter=status_filter,
                           priority_filter=priority_filter,
                           type_filter=type_filter,
                           search=search)


@app.route('/work_orders/dashboard')
def work_order_dashboard():
    """作業工程進捗ダッシュボード"""
    active_orders = WorkOrder.query.filter_by(archived=False).all()
    archived_count = WorkOrder.query.filter_by(archived=True).count()

    # ステータス集計
    status_counts = {
        'pending': sum(1 for o in active_orders if o.status == 'pending'),
        'in_progress': sum(1 for o in active_orders if o.status == 'in_progress'),
        'completed': sum(1 for o in active_orders if o.status == 'completed'),
    }

    # 優先度集計
    priority_counts = {
        'critical': sum(1 for o in active_orders if o.priority == 'critical'),
        'high': sum(1 for o in active_orders if o.priority == 'high'),
        'low': sum(1 for o in active_orders if o.priority == 'low'),
        'none': sum(1 for o in active_orders if o.priority == 'none'),
    }

    # 期限切れ
    overdue = [o for o in active_orders if o.is_overdue()]

    # 今週完了予定
    from datetime import timedelta
    week_later = date.today() + timedelta(days=7)
    due_this_week = [
        o for o in active_orders
        if o.planned_end and date.today() <= o.planned_end <= week_later
        and o.status != 'completed'
    ]

    # 直近完了（7日以内）
    recently_completed = [
        o for o in active_orders
        if o.status == 'completed' and o.actual_end
        and (date.today() - o.actual_end).days <= 7
    ]

    return render_template('work_order_dashboard.html',
                           active_orders=active_orders,
                           archived_count=archived_count,
                           status_counts=status_counts,
                           priority_counts=priority_counts,
                           overdue=overdue,
                           due_this_week=due_this_week,
                           recently_completed=recently_completed)


@app.route('/work_orders/add', methods=['GET', 'POST'])
def add_work_order():
    """作業工程を追加"""
    materials = RawMaterial.query.order_by(RawMaterial.name).all()
    contacts = Contact.query.order_by(Contact.name).all()
    if request.method == 'POST':
        try:
            def parse_date(key):
                val = request.form.get(key, '').strip()
                return datetime.strptime(val, '%Y-%m-%d').date() if val else None

            wo = WorkOrder(
                process_name=request.form.get('process_name', '').strip(),
                lot_name=request.form.get('lot_name', '').strip() or None,
                material_id=int(request.form.get('material_id')) if request.form.get('material_id') else None,
                status=request.form.get('status', 'pending'),
                priority=request.form.get('priority', 'none'),
                invoice_issued=request.form.get('invoice_issued') == '1',
                worker_name=request.form.get('worker_name', '').strip() or None,
                experiment_type=request.form.get('experiment_type', 'standard'),
                planned_start=parse_date('planned_start'),
                planned_end=parse_date('planned_end'),
                notes=request.form.get('notes', '').strip() or None,
            )
            db.session.add(wo)
            db.session.commit()
            flash(f'作業工程「{wo.process_name}」を登録しました', 'success')
            return redirect(url_for('work_orders'))
        except Exception as e:
            db.session.rollback()
            flash(f'登録に失敗しました: {str(e)}', 'danger')

    return render_template('add_work_order.html', materials=materials, contacts=contacts)


@app.route('/work_orders/edit/<int:id>', methods=['GET', 'POST'])
def edit_work_order(id):
    """作業工程を編集"""
    wo = WorkOrder.query.get_or_404(id)
    materials = RawMaterial.query.order_by(RawMaterial.name).all()
    contacts = Contact.query.order_by(Contact.name).all()
    if request.method == 'POST':
        try:
            def parse_date(key):
                val = request.form.get(key, '').strip()
                return datetime.strptime(val, '%Y-%m-%d').date() if val else None

            wo.process_name = request.form.get('process_name', '').strip()
            wo.lot_name = request.form.get('lot_name', '').strip() or None
            wo.material_id = int(request.form.get('material_id')) if request.form.get('material_id') else None
            wo.status = request.form.get('status', 'pending')
            wo.priority = request.form.get('priority', 'none')
            wo.invoice_issued = request.form.get('invoice_issued') == '1'
            wo.worker_name = request.form.get('worker_name', '').strip() or None
            wo.experiment_type = request.form.get('experiment_type', 'standard')
            wo.planned_start = parse_date('planned_start')
            wo.planned_end = parse_date('planned_end')
            wo.actual_start = parse_date('actual_start')
            wo.actual_end = parse_date('actual_end')
            wo.notes = request.form.get('notes', '').strip() or None
            # 実際の日付からステータスを自動更新
            wo.status = wo.compute_status()
            db.session.commit()
            flash(f'作業工程「{wo.process_name}」を更新しました', 'success')
            return redirect(url_for('work_orders'))
        except Exception as e:
            db.session.rollback()
            flash(f'更新に失敗しました: {str(e)}', 'danger')

    return render_template('edit_work_order.html', wo=wo, materials=materials, contacts=contacts)


@app.route('/work_orders/delete/<int:id>', methods=['POST'])
def delete_work_order(id):
    """作業工程を削除"""
    wo = WorkOrder.query.get_or_404(id)
    name = wo.process_name
    db.session.delete(wo)
    db.session.commit()
    flash(f'作業工程「{name}」を削除しました', 'success')
    return redirect(url_for('work_orders'))


@app.route('/work_orders/archive/<int:id>', methods=['POST'])
def archive_work_order(id):
    """作業工程をアーカイブ"""
    wo = WorkOrder.query.get_or_404(id)
    wo.archived = True
    db.session.commit()
    flash(f'「{wo.process_name}」をアーカイブしました', 'success')
    return redirect(url_for('work_orders'))


@app.route('/work_orders/archive_completed', methods=['POST'])
def archive_completed_work_orders():
    """完了済み作業を一括アーカイブ"""
    completed = WorkOrder.query.filter_by(status='completed', archived=False).all()
    count = len(completed)
    for wo in completed:
        wo.archived = True
    if count > 0:
        db.session.commit()
        flash(f'{count}件の完了済み作業をアーカイブしました', 'success')
    else:
        flash('アーカイブ対象の完了済み作業はありません', 'info')
    return redirect(url_for('work_orders'))


@app.route('/work_orders/unarchive/<int:id>', methods=['POST'])
def unarchive_work_order(id):
    """アーカイブを解除"""
    wo = WorkOrder.query.get_or_404(id)
    wo.archived = False
    db.session.commit()
    flash(f'「{wo.process_name}」のアーカイブを解除しました', 'success')
    return redirect(url_for('work_orders', archived='1'))


@app.route('/work_orders/update_status/<int:id>', methods=['POST'])
@csrf.exempt
def update_work_order_status(id):
    """ステータスをAJAXで更新"""
    wo = WorkOrder.query.get_or_404(id)
    data = request.get_json()
    new_status = data.get('status')
    if new_status in ('pending', 'in_progress', 'completed'):
        wo.status = new_status
        if new_status == 'in_progress' and not wo.actual_start:
            wo.actual_start = date.today()
        if new_status == 'completed' and not wo.actual_end:
            wo.actual_end = date.today()
        db.session.commit()
        return jsonify({'ok': True, 'status': wo.status, 'status_label': wo.status_label()})
    return jsonify({'ok': False}), 400


@app.route('/work_orders/calendar')
def work_order_calendar():
    """作業カレンダー（誰がどの作業をするか可視化）"""
    contacts = Contact.query.order_by(Contact.name).all()
    return render_template('work_order_calendar.html', contacts=contacts)


@app.route('/api/work_orders')
def api_work_orders():
    """作業工程データをJSON形式で返す（ダッシュボード・カレンダー用）"""
    active = WorkOrder.query.filter_by(archived=False).all()
    result = []
    for wo in active:
        result.append({
            'id': wo.id,
            'process_name': wo.process_name,
            'lot_name': wo.lot_name,
            'worker_name': wo.worker_name,
            'status': wo.status,
            'status_label': wo.status_label(),
            'priority': wo.priority,
            'priority_label': wo.priority_label(),
            'invoice_issued': wo.invoice_issued,
            'experiment_type': wo.experiment_type,
            'planned_start': wo.planned_start.isoformat() if wo.planned_start else None,
            'planned_end': wo.planned_end.isoformat() if wo.planned_end else None,
            'actual_start': wo.actual_start.isoformat() if wo.actual_start else None,
            'actual_end': wo.actual_end.isoformat() if wo.actual_end else None,
            'is_overdue': wo.is_overdue(),
            'progress_pct': wo.progress_pct(),
            'material_name': wo.material.name if wo.material else None,
        })
    return jsonify(result)


def open_browser():
    """ブラウザを自動的に開く"""
    webbrowser.open('http://127.0.0.1:5000/')

if __name__ == '__main__':
    with app.app_context():
        # 接続プールをクリアしてメタデータをリフレッシュ
        db.engine.dispose()
        
        # データベースとテーブルの作成
        db.create_all()
        
        # 既存テーブルのメタデータを再読み込み
        db.metadata.reflect(bind=db.engine)
        
    # 少し遅延してブラウザを開く
    threading.Timer(1.5, open_browser).start()
    
    # 実行ファイル化されているかチェック
    is_frozen = getattr(sys, 'frozen', False)
    
    # 実行ファイルではdebug=False、開発時はdebug=True
    app.run(debug=not is_frozen, host='127.0.0.1', port=5000)
